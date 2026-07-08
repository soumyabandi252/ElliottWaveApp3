# ══════════════════════════════════════════════════════════════════
# ELLIOTT WAVE MASTER WORKBOOK SCREENER — NASDAQ Composite
#
# Elliott Wave scan of all NASDAQ-listed common stocks (~3,300 tickers).
# Scans all constituents of the NASDAQ Composite index.
#
# OUTPUT:  Elliott_Wave_NASDAQ_Composite_Master_Workbook.xlsx
# INSTALL: pip install yfinance pandas numpy openpyxl requests
# RUN:     python elliott_wave_NASDAQ_Composite.py
#
# NOTE: NASDAQ Composite is a large universe (~3,300 tickers).
#       A full scan can take 3-6 hours. Run overnight or on a server.
# CHANGE ONLY:
#   SCRIPT_DIR = where to save the output file
# ══════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════
# ELLIOTT WAVE MASTER WORKBOOK SCREENER — CUSTOM TICKER LIST
#
# OUTPUT:
#   Elliott_Wave_CUSTOM_Master_Workbook.xlsx
#   ├── Cheat_Sheet
#   ├── Tomorrow_Buys
#   ├── LongTerm_Buys
#   ├── MidTerm_Buys
#   ├── ShortTerm_Buys
#   ├── LongTerm
#   ├── MidTerm
#   ├── ShortTerm
#   └── All_3_Bullish
#
# INSTALL:
#   pip install yfinance pandas numpy openpyxl
#
# RUN:
#   python elliott_wave_custom_tickers.py
#
# CHANGE ONLY THESE INPUTS (see USER INPUTS section below):
#   CUSTOM_TICKERS = ["AAPL", "TSLA", ...]
#   CUSTOM_UNIVERSE_NAME = "My Watchlist"
#   OUTPUT_FILENAME = None
# ══════════════════════════════════════════════════════════════════

import requests
import yfinance as yf
import pandas as pd
import numpy as np
from io import StringIO
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

import json
import os
import yfinance as yf

FUNDAMENTALS_FILE = "fundamentals_history.json"

_FUNDAMENTALS_HISTORY_CACHE = None
_FUNDAMENTALS_HISTORY_DIRTY = False

def _load_fundamentals_history():
    global _FUNDAMENTALS_HISTORY_CACHE
    if _FUNDAMENTALS_HISTORY_CACHE is None:
        history = {}
        if os.path.exists(FUNDAMENTALS_FILE):
            with open(FUNDAMENTALS_FILE, "r") as f:
                try:
                    history = json.load(f)
                except:
                    pass
        _FUNDAMENTALS_HISTORY_CACHE = history
    return _FUNDAMENTALS_HISTORY_CACHE

def _flush_fundamentals_history():
    global _FUNDAMENTALS_HISTORY_DIRTY
    if _FUNDAMENTALS_HISTORY_DIRTY and _FUNDAMENTALS_HISTORY_CACHE is not None:
        with open(FUNDAMENTALS_FILE, "w") as f:
            json.dump(_FUNDAMENTALS_HISTORY_CACHE, f, indent=4)
        _FUNDAMENTALS_HISTORY_DIRTY = False

def get_fundamental_status_with_history(ticker):
    global _FUNDAMENTALS_HISTORY_DIRTY
    try:
        tkr = yf.Ticker(ticker)
        info = tkr.info
        roe = info.get('returnOnEquity', 0)
        pm = info.get('profitMargins', 0)
        rg = info.get('revenueGrowth', 0)

        roe = roe if roe is not None else 0
        pm = pm if pm is not None else 0
        rg = rg if rg is not None else 0

        is_strong = (roe > 0) and (pm > 0) and (rg > 0)
        current_status = "Yes" if is_strong else "No"
    except:
        current_status = "Unknown"

    history = _load_fundamentals_history()

    previous_status = history.get(ticker, current_status)

    if current_status != previous_status and current_status != "Unknown":
        history[ticker] = current_status
        _FUNDAMENTALS_HISTORY_DIRTY = True
        _flush_fundamentals_history()
        return f"{current_status} (Changed from {previous_status})"

    if current_status != "Unknown":
        history[ticker] = current_status
        _FUNDAMENTALS_HISTORY_DIRTY = True
        _flush_fundamentals_history()

    return current_status

warnings.filterwarnings("ignore")

END_DATE = datetime.today()

# ────────────────────────────────────────────────────────────────
# USER INPUTS — CHANGE ONLY THESE
# ────────────────────────────────────────────────────────────────

# ── ADD YOUR TICKERS HERE ────────────────────────────────────────
CUSTOM_TICKERS = [
    "AAPL", "TSLA", "NVDA", "MSFT", "AMZN",
    "GOOGL", "META", "DELL", "AMD", "NFLX",  
# ── BROAD MARKET EQUITY INDICES ─────────────────────────────
    "SP500", "NASDAQ100", "DOW30", "RUSSELL2000", "TOTAL_MARKET", 
    "MSCI_EAFE", "MSCI_EM", "FTSE_DEVELOPED", "VT",
    
    # ── CORE SECTORAL INDICES (S&P SPDR SERIES) ──────────────────
    "XLK", "XLF", "XLV", "XLE", "XLY", "XLP", "XLI", "XLU", "XLB", "XLRE", "XLC",
    
    # ── DEFENSE & AEROSPACE INDICES ─────────────────────────────
    "ITA", "PPA", "XAR", "SHLD", "FITE", "ARKX", "DFEN", "NATO",
    
    # ── SEMICONDUCTORS, AI & ADVANCED TECH ──────────────────────
    "SOXX", "SMH", "BOTZ", "ROBO", "BUG", "CIBR", "SKYY", "WCLD", "FDN", "IBUY",
    
    # ── GREEN ENERGY, COMMODITIES & METALS ──────────────────────
    "GSCI", "LIT", "GLD", "SLV", "USO", "UNG", "BJO", "DBA", "TAN", "FAN", "ICLN", "URA",
    
    # ── FIXED INCOME, TREASURIES & CASH INDICES ──────────────────
    "AGG", "TLT", "IEF", "SHY", "LQD", "HYG", "BNDX", "EMB", "BIL",
    
    # ── GEOPOLITICAL, COUNTRY & REGIONAL INDICES ────────────────
    "FXI", "MCHI", "EWG", "EWJ", "EWU", "INDA", "RSX", "EWW", "EWA", "EWC","ADUR","ASTS","AAPL", "TSLA", "NFLX", "FISV", "JD",   "PYPL", "STNG", "MRNA",
    "NVDA", "MSFT", "AMZN", "BABA", "FNUC", "UPS",  "TGT",  "UNH",
    "NKE",  "SOFI", "AVGO", "GRAB", "PFE","SPX","QQQ","NASDAQ","NIFTY50","GRRR","GNOM","NTRA","LLY","GH","ILMN"
]

# ── OPTIONAL: give your list a name ─────────────────────────────
CUSTOM_UNIVERSE_NAME = "My Watchlist"   # shown in Excel tab headers
CUSTOM_UNIVERSE_CODE = "CUSTOM"         # used in output filename

#OUTPUT_FILENAME = None   # None = auto → Elliott_Wave_CUSTOM_Master_Workbook.xlsx
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILENAME = os.path.join(SCRIPT_DIR, "Elliott_Wave_NASDAQ_Composite_Master_Workbook.xlsx")
# Pivot settings
PIVOT_LEFT_BARS = 5
PIVOT_RIGHT_BARS = 5
PIVOT_NEAR_BARS = 2   # confirmed pivot in last 2 bars or nearby

HEADERS = {
    "User-Agent": "ElliottWaveIndexScreener/1.0 (contact: local-script; educational use)"
}


# ────────────────────────────────────────────────────────────────
# UNIVERSE FETCH HELPERS
# ────────────────────────────────────────────────────────────────
def normalize_ticker_series(s):
    return s.astype(str).str.replace(".", "-", regex=False).str.strip()

def fetch_wikipedia_api_tables(page_title):
    api_url = (
        "https://en.wikipedia.org/w/api.php"
        f"?action=parse&page={page_title}&prop=text&formatversion=2&format=json"
    )
    r = requests.get(api_url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    data = r.json()
    html = data.get("parse", {}).get("text", "")
    if not html:
        raise RuntimeError(f"Wikimedia API returned no parse text for {page_title}.")
    return pd.read_html(StringIO(html))

def fetch_html_tables(url):
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()
    return pd.read_html(StringIO(r.text))


# ────────────────────────────────────────────────────────────────
# S&P 500 AUTO-FETCH
# ────────────────────────────────────────────────────────────────
def get_sp500_constituents():
    api_url = (
        "https://en.wikipedia.org/w/api.php"
        "?action=parse"
        "&page=List_of_S%26P_500_companies"
        "&prop=text"
        "&formatversion=2"
        "&format=json"
    )

    direct_url = "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies"

    last_error = None

    try:
        r = requests.get(api_url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        data = r.json()

        html = data.get("parse", {}).get("text", "")
        if html:
            tables = pd.read_html(StringIO(html))
            for df in tables:
                if "Symbol" in df.columns:
                    sp500 = df.copy()
                    break
            else:
                raise RuntimeError("API HTML returned, but no table with 'Symbol' column found.")
        else:
            raise RuntimeError("Wikimedia API returned no parse text.")

        sp500 = sp500.rename(columns={
            "Symbol": "Ticker",
            "Security": "Company",
            "GICS Sector": "Sector",
            "GICS Sub-Industry": "SubIndustry"
        })

        sp500["Ticker"] = normalize_ticker_series(sp500["Ticker"])

        keep_cols = [c for c in ["Ticker", "Company", "Sector", "SubIndustry"] if c in sp500.columns]
        sp500 = sp500[keep_cols].dropna(subset=["Ticker"]).reset_index(drop=True)

        if len(sp500) < 400:
            raise RuntimeError(f"Parsed too few constituents from API source: {len(sp500)}")

        return sp500

    except Exception as e:
        last_error = f"API method failed: {e}"

    try:
        r = requests.get(direct_url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        tables = pd.read_html(StringIO(r.text))

        sp500 = None
        for df in tables:
            if "Symbol" in df.columns:
                sp500 = df.copy()
                break

        if sp500 is None:
            raise RuntimeError("Direct page returned, but no table with 'Symbol' column found.")

        sp500 = sp500.rename(columns={
            "Symbol": "Ticker",
            "Security": "Company",
            "GICS Sector": "Sector",
            "GICS Sub-Industry": "SubIndustry"
        })

        sp500["Ticker"] = normalize_ticker_series(sp500["Ticker"])

        keep_cols = [c for c in ["Ticker", "Company", "Sector", "SubIndustry"] if c in sp500.columns]
        sp500 = sp500[keep_cols].dropna(subset=["Ticker"]).reset_index(drop=True)

        if len(sp500) < 400:
            raise RuntimeError(f"Parsed too few constituents from direct source: {len(sp500)}")

        return sp500

    except Exception as e:
        last_error = f"{last_error} | Direct method failed: {e}"

    raise RuntimeError(f"Unable to fetch S&P 500 constituents. {last_error}")


# ────────────────────────────────────────────────────────────────
# NASDAQ-100
# ────────────────────────────────────────────────────────────────
def get_nasdaq100_constituents():
    tables = fetch_html_tables("https://en.wikipedia.org/wiki/Nasdaq-100")

    for df in tables:
        cols = [str(c) for c in df.columns]
        if "Ticker" in cols and ("Company" in cols or "Name" in cols):
            out = df.copy()
            if "Name" in out.columns and "Company" not in out.columns:
                out = out.rename(columns={"Name": "Company"})
            if "Sector" not in out.columns:
                out["Sector"] = out["GICS Sector"] if "GICS Sector" in out.columns else ""
            if "SubIndustry" not in out.columns:
                out["SubIndustry"] = out["GICS Sub-Industry"] if "GICS Sub-Industry" in out.columns else ""
            out["Ticker"] = normalize_ticker_series(out["Ticker"])
            return out[["Ticker", "Company", "Sector", "SubIndustry"]].dropna(subset=["Ticker"]).reset_index(drop=True)

    raise RuntimeError("Nasdaq-100 table not found.")


# ────────────────────────────────────────────────────────────────
# DOW 30
# ────────────────────────────────────────────────────────────────
def get_dow30_constituents():
    tables = fetch_html_tables("https://en.wikipedia.org/wiki/Dow_Jones_Industrial_Average")

    for df in tables:
        cols = [str(c) for c in df.columns]
        if "Symbol" in cols and "Company" in cols:
            out = df.rename(columns={"Symbol": "Ticker"}).copy()
            if "Sector" not in out.columns:
                out["Sector"] = ""
            if "SubIndustry" not in out.columns:
                out["SubIndustry"] = ""
            out["Ticker"] = normalize_ticker_series(out["Ticker"])
            return out[["Ticker", "Company", "Sector", "SubIndustry"]].dropna(subset=["Ticker"]).reset_index(drop=True)

    raise RuntimeError("Dow 30 table not found.")


# ────────────────────────────────────────────────────────────────
# RUSSELL 1000
# ────────────────────────────────────────────────────────────────
def get_russell1000_constituents():
    tables = fetch_html_tables("https://en.wikipedia.org/wiki/Russell_1000_Index")

    for df in tables:
        cols = [str(c) for c in df.columns]
        if "Company" in cols and "Ticker" in cols:
            out = df.copy()
            if "Sector" not in out.columns:
                out["Sector"] = ""
            if "SubIndustry" not in out.columns:
                out["SubIndustry"] = ""
            out["Ticker"] = normalize_ticker_series(out["Ticker"])
            return out[["Ticker", "Company", "Sector", "SubIndustry"]].dropna(subset=["Ticker"]).reset_index(drop=True)

    raise RuntimeError("Russell 1000 table not found.")


# ────────────────────────────────────────────────────────────────
# S&P 400
# ────────────────────────────────────────────────────────────────
def get_sp400_constituents():
    tables = fetch_html_tables("https://en.wikipedia.org/wiki/S%26P_400")

    for df in tables:
        cols = [str(c) for c in df.columns]
        if "Symbol" in cols and "Security" in cols:
            out = df.rename(columns={
                "Symbol": "Ticker",
                "Security": "Company",
                "GICS Sector": "Sector",
                "GICS Sub-Industry": "SubIndustry"
            }).copy()
            out["Ticker"] = normalize_ticker_series(out["Ticker"])
            return out[["Ticker", "Company", "Sector", "SubIndustry"]].dropna(subset=["Ticker"]).reset_index(drop=True)

    raise RuntimeError("S&P 400 table not found.")


# ────────────────────────────────────────────────────────────────
# S&P 600
# ────────────────────────────────────────────────────────────────
def get_sp600_constituents():
    tables = fetch_html_tables("https://en.wikipedia.org/wiki/S%26P_600")

    for df in tables:
        cols = [str(c) for c in df.columns]
        if "Symbol" in cols and ("Company" in cols or "Security" in cols):
            out = df.copy()
            if "Security" in out.columns and "Company" not in out.columns:
                out = out.rename(columns={"Security": "Company"})
            out = out.rename(columns={"Symbol": "Ticker"})
            if "Sector" not in out.columns:
                out["Sector"] = out["GICS Sector"] if "GICS Sector" in out.columns else ""
            if "SubIndustry" not in out.columns:
                out["SubIndustry"] = out["GICS Sub-Industry"] if "GICS Sub-Industry" in out.columns else ""
            out["Ticker"] = normalize_ticker_series(out["Ticker"])
            return out[["Ticker", "Company", "Sector", "SubIndustry"]].dropna(subset=["Ticker"]).reset_index(drop=True)

    raise RuntimeError("S&P 600 table not found.")


# ────────────────────────────────────────────────────────────────
# NASDAQ BROAD LIST
# ────────────────────────────────────────────────────────────────
def get_nasdaq_listed_common():
    url = "https://www.nasdaqtrader.com/dynamic/SymDir/nasdaqlisted.txt"
    r = requests.get(url, headers=HEADERS, timeout=30)
    r.raise_for_status()

    lines = r.text.splitlines()
    rows = [line.split("|") for line in lines if "|" in line]
    df = pd.DataFrame(rows[1:], columns=rows[0])

    df = df[df["Symbol"] != "File Creation Time"].copy()
    df = df[df["ETF"] == "N"].copy()
    df["Ticker"] = normalize_ticker_series(df["Symbol"])
    df["Company"] = df["Security Name"]
    df["Sector"] = ""
    df["SubIndustry"] = ""

    return df[["Ticker", "Company", "Sector", "SubIndustry"]].dropna(subset=["Ticker"]).reset_index(drop=True)


def get_nasdaq_composite_proxy():
    return get_nasdaq_listed_common()


def get_nasdaq1000_proxy():
    base = get_nasdaq_listed_common()
    tickers = base["Ticker"].tolist()
    caps = [None] * len(tickers)

    def _fetch_cap(idx_ticker):
        idx, ticker = idx_ticker
        try:
            info = yf.Ticker(ticker).fast_info
            return idx, ticker, info.get("marketCap", None)
        except Exception:
            return idx, ticker, None

    done_count = 0
    with ThreadPoolExecutor(max_workers=16) as executor:
        futures = [executor.submit(_fetch_cap, (i, t)) for i, t in enumerate(tickers)]
        for future in as_completed(futures):
            idx, ticker, cap = future.result()
            caps[idx] = (ticker, cap)
            done_count += 1
            if done_count % 250 == 0:
                print(f"Ranking Nasdaq universe by market cap: {done_count}/{len(tickers)}", end="\r")

    cap_df = pd.DataFrame(caps, columns=["Ticker", "MarketCap"]).dropna()
    top = cap_df.sort_values("MarketCap", ascending=False).head(1000)[["Ticker"]]
    out = base.merge(top, on="Ticker", how="inner").reset_index(drop=True)
    return out


# ────────────────────────────────────────────────────────────────
# RUSSELL 2000 / IWM PROXY
# ────────────────────────────────────────────────────────────────
def get_russell2000_proxy():
    return get_sp600_constituents()

def get_iwm_proxy():
    return get_russell2000_proxy()


# ────────────────────────────────────────────────────────────────
# TOTAL U.S. PROXY
# ────────────────────────────────────────────────────────────────
def get_total_us_proxy():
    sp500 = get_sp500_constituents()
    sp400 = get_sp400_constituents()
    sp600 = get_sp600_constituents()
    nasdaq = get_nasdaq_listed_common()

    out = pd.concat([sp500, sp400, sp600, nasdaq], ignore_index=True)
    out = out.drop_duplicates(subset=["Ticker"]).reset_index(drop=True)
    return out


# ────────────────────────────────────────────────────────────────
# UNIVERSE ROUTER
# ────────────────────────────────────────────────────────────────
def get_universe(choice):
    choice = str(choice).upper().strip()

    if choice == "SP500":
        return "S&P 500", "SP500", get_sp500_constituents()
    elif choice == "NASDAQ100":
        return "Nasdaq-100", "NASDAQ100", get_nasdaq100_constituents()
    elif choice == "DOW30":
        return "Dow 30", "DOW30", get_dow30_constituents()
    elif choice == "RUSSELL1000":
        return "Russell 1000", "RUSSELL1000", get_russell1000_constituents()
    elif choice == "RUSSELL2000":
        return "Russell 2000 Proxy", "RUSSELL2000", get_russell2000_proxy()
    elif choice == "IWM":
        return "IWM / Russell 2000 Proxy", "IWM", get_iwm_proxy()
    elif choice == "SP400":
        return "S&P 400", "SP400", get_sp400_constituents()
    elif choice == "SP600":
        return "S&P 600", "SP600", get_sp600_constituents()
    elif choice == "NASDAQ_COMPOSITE":
        return "Nasdaq Composite Proxy", "NASDAQ_COMPOSITE", get_nasdaq_composite_proxy()
    elif choice == "NASDAQ1000":
        return "Nasdaq-1000 Proxy", "NASDAQ1000", get_nasdaq1000_proxy()
    elif choice == "TOTAL_US":
        return "Total U.S. Proxy", "TOTAL_US", get_total_us_proxy()
    else:
        raise ValueError(
            "Unsupported INDEX_CHOICE. Use one of: "
            "SP500, NASDAQ100, DOW30, RUSSELL1000, RUSSELL2000, IWM, "
            "SP400, SP600, NASDAQ_COMPOSITE, NASDAQ1000, TOTAL_US"
        )


# ────────────────────────────────────────────────────────────────
# DATA DOWNLOAD
# ────────────────────────────────────────────────────────────────
def download_data(ticker, days, interval="1d"):
    start = END_DATE - timedelta(days=days)
    try:
        df = yf.download(
            ticker,
            start=start,
            end=END_DATE,
            interval=interval,
            progress=False,
            auto_adjust=True,
            threads=False
        )

        if df.empty or len(df) < 30:
            return None

        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
        needed = ["Open", "High", "Low", "Close", "Volume"]

        if not all(col in df.columns for col in needed):
            return None

        return df[needed].dropna()

    except Exception:
        return None




def download_data_monthly(ticker, years=10):
    """Download monthly OHLCV data for ultra-long-term EW analysis."""
    start = END_DATE - timedelta(days=365 * years)
    try:
        df = yf.download(
            ticker,
            start=start,
            end=END_DATE,
            interval="1mo",
            progress=False,
            auto_adjust=True,
            threads=False
        )
        if df.empty or len(df) < 12:
            return None
        df.columns = [c[0] if isinstance(c, tuple) else c for c in df.columns]
        needed = ["Open", "High", "Low", "Close", "Volume"]
        if not all(col in df.columns for col in needed):
            return None
        return df[needed].dropna()
    except Exception:
        return None

# ────────────────────────────────────────────────────────────────
# INDICATORS
# ────────────────────────────────────────────────────────────────
def ema(arr, period):
    return pd.Series(arr).ewm(span=period, adjust=False).mean().values

def rsi_calc(arr, period=14):
    delta = np.diff(arr)
    gains = np.where(delta > 0, delta, 0)
    losses = np.where(delta < 0, -delta, 0)
    avg_gain = pd.Series(gains).ewm(alpha=1/period, adjust=False).mean().values  # Wilder RMA
    avg_loss = pd.Series(losses).ewm(alpha=1/period, adjust=False).mean().values  # Wilder RMA
    rs = avg_gain / (avg_loss + 1e-10)
    return 100 - 100 / (1 + rs)

def macd_calc(arr, fast=12, slow=26, signal=9):
    fast_ema = pd.Series(arr).ewm(span=fast, adjust=False).mean()
    slow_ema = pd.Series(arr).ewm(span=slow, adjust=False).mean()
    macd_line = fast_ema - slow_ema
    signal_line = macd_line.ewm(span=signal, adjust=False).mean()
    hist = macd_line - signal_line
    return macd_line.values, signal_line.values, hist.values

def obv_calc(close, volume):
    close = np.asarray(close, dtype=float)
    volume = np.asarray(volume, dtype=float)
    if len(close) == 0:
        return np.array([])
    direction = np.sign(np.diff(close))
    signed_vol = direction * volume[1:]
    obv = np.concatenate(([0.0], np.cumsum(signed_vol)))
    return obv

def fmt_price(x):
    try:
        return round(float(x), 2)
    except Exception:
        return None


# ────────────────────────────────────────────────────────────────
# PIVOT DETECTION (Pine-style confirmed pivots)
# ────────────────────────────────────────────────────────────────
def detect_confirmed_pivots(df, left_bars=5, right_bars=5):
    if df is None or len(df) < left_bars + right_bars + 3:
        return pd.DataFrame(columns=[
            "PivotType", "PivotBar", "ConfirmBar",
            "PivotDate", "ConfirmDate", "PivotPrice"
        ])

    highs = df["High"].values
    lows = df["Low"].values
    idx = df.index
    n = len(df)
    win = left_bars + right_bars + 1

    from numpy.lib.stride_tricks import sliding_window_view
    hi_windows = sliding_window_view(highs, win)
    lo_windows = sliding_window_view(lows, win)

    center_idx = np.arange(left_bars, n - right_bars)
    center_highs = highs[center_idx]
    center_lows = lows[center_idx]

    is_pivot_high = (center_highs == hi_windows.max(axis=1)) & \
                    (np.sum(hi_windows == center_highs[:, None], axis=1) == 1)
    is_pivot_low = (center_lows == lo_windows.min(axis=1)) & \
                   (np.sum(lo_windows == center_lows[:, None], axis=1) == 1)

    confirm_idx = center_idx + right_bars
    rows = []

    hi_positions = center_idx[is_pivot_high]
    hi_confirms = confirm_idx[is_pivot_high]
    for i, ci in zip(hi_positions, hi_confirms):
        rows.append({
            "PivotType": "HIGH",
            "PivotBar": int(i),
            "ConfirmBar": int(ci),
            "PivotDate": idx[i],
            "ConfirmDate": idx[ci],
            "PivotPrice": float(highs[i]),
        })

    lo_positions = center_idx[is_pivot_low]
    lo_confirms = confirm_idx[is_pivot_low]
    for i, ci in zip(lo_positions, lo_confirms):
        rows.append({
            "PivotType": "LOW",
            "PivotBar": int(i),
            "ConfirmBar": int(ci),
            "PivotDate": idx[i],
            "ConfirmDate": idx[ci],
            "PivotPrice": float(lows[i]),
        })

    return pd.DataFrame(rows)


def recent_pivot_summary(df, left_bars=5, right_bars=5, near_bars=2):
    out = {
        "Pivot_Low": "",
        "Pivot_Low_Date": "",
        "Pivot_Low_Price": "",
        "Pivot_Low_Bars_Ago": "",
        "Pivot_High": "",
        "Pivot_High_Date": "",
        "Pivot_High_Price": "",
        "Pivot_High_Bars_Ago": "",
    }

    piv = detect_confirmed_pivots(df, left_bars=left_bars, right_bars=right_bars)
    if piv.empty:
        return out

    last_bar_num = len(df) - 1

    lows = piv[piv["PivotType"] == "LOW"].copy()
    highs = piv[piv["PivotType"] == "HIGH"].copy()

    if not lows.empty:
        lows["BarsAgoFromConfirmation"] = last_bar_num - lows["ConfirmBar"]
        recent_lows = lows[lows["BarsAgoFromConfirmation"] <= near_bars].sort_values("ConfirmBar", ascending=False)
        if not recent_lows.empty:
            r = recent_lows.iloc[0]
            out["Pivot_Low"] = "YES"
            out["Pivot_Low_Date"] = str(pd.to_datetime(r["PivotDate"]).date())
            out["Pivot_Low_Price"] = round(float(r["PivotPrice"]), 2)
            out["Pivot_Low_Bars_Ago"] = int(r["BarsAgoFromConfirmation"])

    if not highs.empty:
        highs["BarsAgoFromConfirmation"] = last_bar_num - highs["ConfirmBar"]
        recent_highs = highs[highs["BarsAgoFromConfirmation"] <= near_bars].sort_values("ConfirmBar", ascending=False)
        if not recent_highs.empty:
            r = recent_highs.iloc[0]
            out["Pivot_High"] = "YES"
            out["Pivot_High_Date"] = str(pd.to_datetime(r["PivotDate"]).date())
            out["Pivot_High_Price"] = round(float(r["PivotPrice"]), 2)
            out["Pivot_High_Bars_Ago"] = int(r["BarsAgoFromConfirmation"])

    return out


# ────────────────────────────────────────────────────────────────
# GUIDANCE
# ────────────────────────────────────────────────────────────────
def wave_guidance(state, close, f382, f500, f618, f786, f100, f161, hi52, lo52, m20, m50, m200):
    s = str(state).upper()

    current_target = None
    ideal_entry = None
    pullback_zone = None
    stop_loss = None
    extension_target = None
    invalidation = None
    if_extended_then = ""
    beginner_note = ""
    trade_quality = ""

    if "WAVE 2" in s:
        current_target = hi52
        ideal_entry = f618
        pullback_zone = f"{fmt_price(f618)} - {fmt_price(f786)}"
        stop_loss = f100
        invalidation = f100
        extension_target = hi52 + (hi52 - lo52) * 0.272
        if_extended_then = (
            f"If price rebounds strongly from this correction, first target is the old high near {fmt_price(hi52)}. "
            f"If momentum expands further, the move can extend toward {fmt_price(extension_target)}."
        )
        beginner_note = (
            "This is one of the best beginner-style entries because you are buying a pullback, not chasing a rally. "
            "Enter near the zone and keep the stop below invalidation."
        )
        trade_quality = "HIGH QUALITY DIP BUY"

    elif "ACCUMULATION" in s or "CAPITULATION" in s:
        current_target = hi52
        ideal_entry = f500 if "ACCUMULATION" in s else close
        pullback_zone = f"{fmt_price(f500)} - {fmt_price(f618)}"
        stop_loss = lo52
        invalidation = lo52
        extension_target = hi52 + (hi52 - lo52) * 0.618
        if_extended_then = (
            f"If the stock truly starts a new cycle, it can reclaim the old high and later extend toward {fmt_price(extension_target)}."
        )
        beginner_note = (
            "This is an early-cycle opportunity. It can be rewarding, but it may not move immediately. "
            "Start small and only add after price confirms strength."
        )
        trade_quality = "EARLY OPPORTUNITY"

    elif "WAVE 3" in s or "3RD OF 3RD" in s or "BREAKOUT" in s:
        current_target = f161 if f161 is not None else hi52
        ideal_entry = m20 if close > m20 else close
        pullback_zone = f"{fmt_price(m20)} - {fmt_price(m50)}"
        stop_loss = m50
        invalidation = m50
        extension_target = hi52 + (hi52 - lo52) * 0.618
        if_extended_then = (
            f"If this Wave 3 becomes extended, price can run beyond the normal target and push toward {fmt_price(extension_target)}."
        )
        beginner_note = (
            "This is usually the strongest trend phase. Beginners should avoid buying after a giant green candle. "
            "Wait for a pullback toward the 20 EMA or 50 EMA."
        )
        trade_quality = "TREND FOLLOWING — VERY GOOD"

    elif "TRIANGLE" in s:
        current_target = hi52
        ideal_entry = close
        pullback_zone = f"{fmt_price(f382)} - {fmt_price(f500)}"
        stop_loss = f618
        invalidation = f618
        extension_target = f161
        if_extended_then = (
            f"If the triangle resolves upward with volume, the thrust can first target {fmt_price(hi52)} and then possibly extend toward {fmt_price(f161)}."
        )
        beginner_note = "Do not force a trade inside a tight range. Wait for breakout confirmation."
        trade_quality = "WAIT FOR CONFIRMATION"

    elif "WAVE 5" in s or "LATE BULL" in s:
        current_target = hi52
        ideal_entry = ""
        pullback_zone = f"{fmt_price(m20)} - {fmt_price(m50)}"
        stop_loss = m50
        invalidation = m50
        extension_target = hi52 + (hi52 - lo52) * 0.272
        if_extended_then = (
            f"Wave 5 can still extend toward {fmt_price(extension_target)}, but the risk/reward is worse and reversals can happen quickly."
        )
        beginner_note = "This is not the safest fresh entry for a beginner. Better to protect profits or wait for the next correction."
        trade_quality = "LOWER QUALITY — DO NOT CHASE"

    elif "BLOWOFF" in s or "TRAP" in s or "BEAR" in s or "EXIT" in s:
        current_target = m50
        ideal_entry = ""
        pullback_zone = ""
        stop_loss = ""
        invalidation = ""
        extension_target = m200
        if_extended_then = "This is not a healthy bullish extension setup. If weakness continues, price may retrace deeper toward major support."
        beginner_note = "Avoid fresh buying here. Let the chart reset and wait for a cleaner setup."
        trade_quality = "AVOID NEW ENTRY"

    else:
        current_target = hi52
        ideal_entry = f500
        pullback_zone = f"{fmt_price(f500)} - {fmt_price(f618)}"
        stop_loss = f786
        invalidation = f100
        extension_target = f161
        if_extended_then = f"If bullish momentum improves, price may move toward {fmt_price(f161)}. If it breaks invalidation, wait for a better setup."
        beginner_note = "This setup is not fully mature yet. Wait for either a clearer pullback or a confirmed breakout."
        trade_quality = "MEDIUM / WAIT"

    return {
        "Current_Wave_Target": fmt_price(current_target) if current_target != "" else "",
        "Ideal_Entry_Price": fmt_price(ideal_entry) if ideal_entry not in ["", None] else "",
        "Pullback_Buy_Zone": pullback_zone,
        "Stop_Loss": fmt_price(stop_loss) if stop_loss not in ["", None] else "",
        "Invalidation_Level": fmt_price(invalidation) if invalidation not in ["", None] else "",
        "Extension_Target": fmt_price(extension_target) if extension_target not in ["", None] else "",
        "If_Extended_Then": if_extended_then,
        "Beginner_Note": beginner_note,
        "Trade_Quality": trade_quality
    }


# ────────────────────────────────────────────────────────────────
# SCREENING ENGINE
# ────────────────────────────────────────────────────────────────

# ════════════════════════════════════════════════════════════════
# EW STATE META — Duration / Next-State / Transition helpers
# ════════════════════════════════════════════════════════════════
_STATE_DURATION_BARS = {
    # Long (weekly bars)
    "CAPITULATION — BULL CYCLE STARTING":  (4,  16),
    "ACCUMULATION — DOW PHASE 1":          (8,  26),
    "WAVE 2 BUY ZONE — BEST LONG ENTRY":   (4,  13),
    "WAVE 3 — STRONGEST WAVE (RIDE IT)":   (13, 52),
    "WAVE 5 — LATE BULL, START TRIMMING":  (4,  13),
    "BLOWOFF TOP — EXIT ALL LONGS":        (1,   4),
    "BEAR MARKET — AVOID / CASH":          (13, 78),
    "B-WAVE BULL TRAP — FAKE RALLY":       (2,   8),
    "TRANSITIONING / SIDEWAYS":            (4,  26),
    # Mid (daily bars)
    "WAVE 2 BUY ZONE (BEST RISK:REWARD)":  (5,  21),
    "WAVE 3 STARTING — ENTER NOW":         (3,   8),
    "WAVE 3 ACTIVE — HOLD / TRAIL":        (10, 40),
    "WAVE 3 ACTIVE":                       (10, 40),
    "TRIANGLE — WAIT FOR THRUST":          (5,  20),
    "WAVE 5 — REDUCE RISK":                (5,  15),
    "B-WAVE TRAP — STAND ASIDE":           (3,  10),
    "ACCUMULATION — BUILD SLOWLY":         (10, 30),
    "WATCH LIST — SETUP FORMING":          (3,  15),
    # Short (daily bars)
    "3RD OF 3RD WAVE — DO NOT SHORT":      (2,   6),
    "WAVE 3 BREAKOUT — ENTER NOW":         (2,   6),
    "TRIANGLE THRUST — BUY BREAKOUT":      (2,   8),
    "WAVE 3 ACTIVE — RIDE IT":             (5,  20),
    "WAVE 5 TOP — START EXITING":          (3,  10),
    "BLOWOFF — EXIT IMMEDIATELY":          (1,   3),
    "SETUP FORMING — WAIT FOR SIGNAL":     (3,  10),
}

_NEXT_STATE_MAP = {
    # Long
    "CAPITULATION — BULL CYCLE STARTING":  ("ACCUMULATION — DOW PHASE 1",          6,  13),
    "ACCUMULATION — DOW PHASE 1":          ("WAVE 2 BUY ZONE — BEST LONG ENTRY",   4,  10),
    "WAVE 2 BUY ZONE — BEST LONG ENTRY":   ("WAVE 3 — STRONGEST WAVE (RIDE IT)",   2,   8),
    "WAVE 3 — STRONGEST WAVE (RIDE IT)":   ("WAVE 5 — LATE BULL, START TRIMMING", 13,  52),
    "WAVE 5 — LATE BULL, START TRIMMING":  ("BLOWOFF TOP — EXIT ALL LONGS",        2,   8),
    "BLOWOFF TOP — EXIT ALL LONGS":        ("BEAR MARKET — AVOID / CASH",          1,   3),
    "BEAR MARKET — AVOID / CASH":          ("CAPITULATION — BULL CYCLE STARTING", 13,  52),
    "B-WAVE BULL TRAP — FAKE RALLY":       ("BEAR MARKET — AVOID / CASH",          2,   8),
    "TRANSITIONING / SIDEWAYS":            ("WAVE 2 BUY ZONE — BEST LONG ENTRY",  4,  13),
    # Mid
    "WAVE 2 BUY ZONE (BEST RISK:REWARD)":  ("WAVE 3 STARTING — ENTER NOW",        3,   8),
    "WAVE 3 STARTING — ENTER NOW":         ("WAVE 3 ACTIVE — HOLD / TRAIL",       1,   3),
    "WAVE 3 ACTIVE — HOLD / TRAIL":        ("WAVE 5 — REDUCE RISK",              10,  40),
    "WAVE 3 ACTIVE":                       ("WAVE 5 — REDUCE RISK",              10,  40),
    "TRIANGLE — WAIT FOR THRUST":          ("WAVE 3 ACTIVE — HOLD / TRAIL",       2,   6),
    "WAVE 5 — REDUCE RISK":                ("B-WAVE TRAP — STAND ASIDE",          5,  15),
    "B-WAVE TRAP — STAND ASIDE":           ("ACCUMULATION — BUILD SLOWLY",        3,  10),
    "ACCUMULATION — BUILD SLOWLY":         ("WAVE 2 BUY ZONE (BEST RISK:REWARD)", 5,  15),
    "WATCH LIST — SETUP FORMING":          ("WAVE 2 BUY ZONE (BEST RISK:REWARD)", 3,  10),
    # Short
    "3RD OF 3RD WAVE — DO NOT SHORT":      ("WAVE 5 TOP — START EXITING",         2,   6),
    "WAVE 3 BREAKOUT — ENTER NOW":         ("WAVE 3 ACTIVE — RIDE IT",            1,   3),
    "TRIANGLE THRUST — BUY BREAKOUT":      ("WAVE 3 ACTIVE — RIDE IT",            1,   3),
    "WAVE 3 ACTIVE — RIDE IT":             ("WAVE 5 TOP — START EXITING",         5,  20),
    "WAVE 5 TOP — START EXITING":          ("BLOWOFF — EXIT IMMEDIATELY",         2,   8),
    "BLOWOFF — EXIT IMMEDIATELY":          ("SETUP FORMING — WAIT FOR SIGNAL",    1,   3),
    "SETUP FORMING — WAIT FOR SIGNAL":     ("WAVE 3 BREAKOUT — ENTER NOW",        3,  10),
}

_TRANSITION_TRIGGER_MAP = {
    # Long
    "CAPITULATION — BULL CYCLE STARTING":  "OBV turns up + RSI > 35 + 2 higher-low weekly closes",
    "ACCUMULATION — DOW PHASE 1":          "Price reclaims 50 EMA with expanding volume",
    "WAVE 2 BUY ZONE — BEST LONG ENTRY":   "Bounce from 61.8% Fib with MACD cross-up",
    "WAVE 3 — STRONGEST WAVE (RIDE IT)":   "All 3 MAs bull-aligned + MACD histogram expanding",
    "WAVE 5 — LATE BULL, START TRIMMING":  "RSI > 70 divergence + OBV bear-div forming",
    "BLOWOFF TOP — EXIT ALL LONGS":        "Volume climax candle + RSI > 80 reversal",
    "BEAR MARKET — AVOID / CASH":          "Price reclaims 200 EMA + MACD cross-up above zero",
    "B-WAVE BULL TRAP — FAKE RALLY":       "MACD cross-down + price breaks back below 200 EMA",
    "TRANSITIONING / SIDEWAYS":            "MACD cross-up or 61.8% Fib bounce with volume",
    # Mid
    "WAVE 2 BUY ZONE (BEST RISK:REWARD)":  "MACD cross-up + price holds 61.8% Fib level",
    "WAVE 3 STARTING — ENTER NOW":         "Price + volume break above recent swing high",
    "WAVE 3 ACTIVE — HOLD / TRAIL":        "Price pulls back to 20 EMA without MACD cross-down",
    "WAVE 3 ACTIVE":                       "Price pulls back to 20 EMA without MACD cross-down",
    "TRIANGLE — WAIT FOR THRUST":          "Volume surge + close above triangle upper boundary",
    "WAVE 5 — REDUCE RISK":                "OBV bear-div + RSI > 65 with MACD histogram shrinking",
    "B-WAVE TRAP — STAND ASIDE":           "Break of most recent swing low on volume",
    "ACCUMULATION — BUILD SLOWLY":         "Price reclaims 50 EMA + OBV trending up",
    "WATCH LIST — SETUP FORMING":          "MACD cross-up + RSI > 50 + price > 20 EMA",
    # Short
    "3RD OF 3RD WAVE — DO NOT SHORT":      "Volume dries up + RSI > 75 divergence",
    "WAVE 3 BREAKOUT — ENTER NOW":         "Sustained close + volume above breakout level",
    "TRIANGLE THRUST — BUY BREAKOUT":      "Second close above triangle boundary on volume",
    "WAVE 3 ACTIVE — RIDE IT":             "20 EMA acts as support + hold while RSI > 50",
    "WAVE 5 TOP — START EXITING":          "Volume spike + bearish engulfing daily candle",
    "BLOWOFF — EXIT IMMEDIATELY":          "Any close below 20 EMA with volume > 1.5x average",
    "SETUP FORMING — WAIT FOR SIGNAL":     "MACD cross-up + RSI > 50 + price > 50 EMA",
}


def _bars_to_human(bars_min, bars_max, horizon):
    """Convert a bar-count range to a readable human string."""
    if horizon == "long":  # weekly bars
        wk_lo, wk_hi = bars_min, bars_max
        if wk_hi <= 4:
            return f"~{wk_lo}-{wk_hi} weeks"
        mo_lo = round(wk_lo / 4.3, 1)
        mo_hi = round(wk_hi / 4.3, 1)
        if mo_hi <= 3:
            return f"~{mo_lo}-{mo_hi} months"
        return f"~{round(mo_lo)}-{round(mo_hi)} months"
    else:  # daily bars (mid + short)
        if bars_max <= 10:
            return f"~{bars_min}-{bars_max} days"
        wk_lo = round(bars_min / 5, 1)
        wk_hi = round(bars_max / 5, 1)
        if wk_hi <= 8:
            return f"~{wk_lo}-{wk_hi} weeks"
        mo_lo = round(wk_lo / 4.3, 1)
        mo_hi = round(wk_hi / 4.3, 1)
        return f"~{mo_lo}-{mo_hi} months"


def _current_state_bars(df, state_key, horizon):
    """Count consecutive bars the stock has been in the current EW state."""
    if df is None or len(df) < 5:
        return 0
    c = df["Close"].values
    n = len(c)
    p200 = min(200, max(20, n // 2))
    p50  = min(50,  max(10, n // 3))
    m200 = pd.Series(c).ewm(span=p200, adjust=False).mean().values
    m50  = pd.Series(c).ewm(span=p50,  adjust=False).mean().values
    s = state_key.upper()
    if "WAVE 3" in s or "3RD OF 3RD" in s or "BREAKOUT" in s:
        cond = lambda i: c[i] > m200[i] and c[i] > m50[i]
    elif "WAVE 2" in s or "BUY ZONE" in s:
        cond = lambda i: c[i] > m200[i]
    elif "CAPITULATION" in s or "ACCUMULATION" in s:
        cond = lambda i: c[i] <= m200[i]
    elif "BEAR" in s or "BLOWOFF" in s or "TRAP" in s:
        cond = lambda i: c[i] < m200[i]
    else:
        cond = lambda i: True
    count = 0
    for i in range(n - 1, max(n - 200, 0), -1):
        if cond(i):
            count += 1
        else:
            break
    return count



# ════════════════════════════════════════════════════════════════
# FROST & PRECHTER — COMPLETE RULE SET ADDITIONS
# ════════════════════════════════════════════════════════════════

# ────────────────────────────────────────────────────────────────
# A1. WAVE DEGREE CLASSIFIER
#     Supercycle / Cycle / Primary / Intermediate / Minor / Minute
# ────────────────────────────────────────────────────────────────
def classify_wave_degree(horizon, n_bars, bar_interval="1wk"):
    """
    Assigns Frost & Prechter wave degree based on timeframe + bar count.
    Grand Supercycle → years/decades (beyond our data)
    Supercycle       → weekly 5yr
    Cycle            → weekly 1-2yr
    Primary          → daily 6-12mo
    Intermediate     → daily 1-3mo
    Minor            → daily 2-6wk
    Minute           → daily <2wk
    """
    if bar_interval == "1wk":
        if n_bars >= 200:
            return "Supercycle"
        elif n_bars >= 80:
            return "Cycle"
        else:
            return "Primary"
    else:  # daily
        if n_bars >= 180:
            return "Primary"
        elif n_bars >= 60:
            return "Intermediate"
        elif n_bars >= 20:
            return "Minor"
        else:
            return "Minute"


# ────────────────────────────────────────────────────────────────
# A2. CORRECTIVE PATTERN CLASSIFIER
#     Zigzag (5-3-5) | Flat (3-3-5) | Triangle (3-3-3-3-3)
#     Double Three | Irregular
# ────────────────────────────────────────────────────────────────
def classify_corrective_pattern(pivots_df, close_arr, swg_hi, swg_lo):
    """
    Uses confirmed pivots to classify the most recent corrective structure.
    Returns a string: Zigzag / Flat / Triangle / Double-Three / Unknown
    """
    if pivots_df is None or pivots_df.empty or len(close_arr) < 20:
        return "Unknown"

    lows_  = pivots_df[pivots_df["PivotType"] == "LOW"].sort_values("PivotBar")
    highs_ = pivots_df[pivots_df["PivotType"] == "HIGH"].sort_values("PivotBar")

    if len(lows_) < 2 or len(highs_) < 1:
        return "Unknown"

    # Most recent A-low and B-high
    last_low_a  = lows_.iloc[-1]["PivotPrice"]
    last_low_b  = lows_.iloc[-2]["PivotPrice"] if len(lows_) >= 2 else None
    last_high_b = highs_.iloc[-1]["PivotPrice"]
    cur         = close_arr[-1]
    fib_range   = swg_hi - swg_lo if swg_hi > swg_lo else 0.001

    # Zigzag: Wave B retraces < 61.8% of Wave A
    if last_low_b is not None:
        wave_a = abs(swg_hi - last_low_a)
        wave_b = abs(last_high_b - last_low_a)
        b_retrace = wave_b / wave_a if wave_a > 0 else 0

        if b_retrace < 0.618:
            return "Zigzag (5-3-5)"
        elif 0.80 <= b_retrace <= 1.02:
            return "Flat (3-3-5)"
        elif b_retrace > 1.02:
            return "Expanded Flat / Irregular"

    # Triangle: range contracting in last 5 pivots
    all_pivots = pivots_df.sort_values("PivotBar").tail(10)
    if len(all_pivots) >= 6:
        prices = all_pivots["PivotPrice"].values
        highs_seq = prices[1::2]
        lows_seq  = prices[::2]
        if len(highs_seq) >= 2 and len(lows_seq) >= 2:
            h_contracting = highs_seq[-1] < highs_seq[-2] if len(highs_seq) >= 2 else False
            l_contracting = lows_seq[-1]  > lows_seq[-2]  if len(lows_seq)  >= 2 else False
            if h_contracting and l_contracting:
                return "Triangle (3-3-3-3-3)"

    return "Simple Correction / Double-Three"


# ────────────────────────────────────────────────────────────────
# A3. DIAGONAL TRIANGLE DETECTOR (Frost Ch 1-2)
#     Ending Diagonal = Wave 5 or Wave C converging wedge
#     Leading Diagonal = Wave 1 or Wave A converging wedge
# ────────────────────────────────────────────────────────────────
def detect_diagonal_triangle(df, horizon="short"):
    """
    Detects wedge / diagonal triangle pattern using recent 20-40 bars.
    Returns: 'Ending Diagonal', 'Leading Diagonal', or None
    """
    if df is None or len(df) < 20:
        return None

    n = min(40, len(df))
    highs  = df["High"].values[-n:]
    lows   = df["Low"].values[-n:]
    closes = df["Close"].values[-n:]

    # Fit linear trend to highs and lows
    x = np.arange(n)
    # Slope of upper boundary (highs)
    hi_slope = np.polyfit(x, highs, 1)[0]
    # Slope of lower boundary (lows)
    lo_slope = np.polyfit(x, lows, 1)[0]

    # Converging wedge: upper slope falling, lower slope rising (or both narrowing)
    converging = hi_slope < 0 and lo_slope > 0
    narrowing  = abs(highs[-1] - lows[-1]) < abs(highs[0] - lows[0]) * 0.75

    if not (converging or narrowing):
        return None

    # Ending diagonal: appears at end of Wave 5 (price near 52w high, RSI diverging)
    rsi = rsi_calc(closes)
    rsi_diverging = len(rsi) > 10 and closes[-1] > closes[-10] and rsi[-1] < rsi[-10]
    if rsi_diverging:
        return "Ending Diagonal (Wave 5 warning)"

    # Leading diagonal: appears at start (price recovering from lows)
    price_recovering = closes[-1] > closes[0]
    if price_recovering and not rsi_diverging:
        return "Leading Diagonal (Wave 1 / Wave A)"

    return "Diagonal Triangle (Unspecified)"


# ────────────────────────────────────────────────────────────────
# A4. TREND CHANNEL BUILDER (Frost Ch 2 — Channeling)
#     Connects W2 low → W4 low (base channel)
#     Projects W5 target from W3 top (parallel)
# ────────────────────────────────────────────────────────────────
def build_wave_channel(pivots_df, df):
    """
    Frost & Prechter channeling technique:
    1) Draw line through W2 and W4 bottoms
    2) Parallel through W3 top → gives W5 target zone
    Returns dict with channel values or empty dict.
    """
    if pivots_df is None or pivots_df.empty or df is None or len(df) < 20:
        return {"Channel_Base": "", "Channel_Top": "", "W5_Channel_Target": ""}

    lows  = pivots_df[pivots_df["PivotType"] == "LOW"].sort_values("PivotBar")
    highs = pivots_df[pivots_df["PivotType"] == "HIGH"].sort_values("PivotBar")

    if len(lows) < 2 or len(highs) < 1:
        return {"Channel_Base": "", "Channel_Top": "", "W5_Channel_Target": ""}

    # Use most recent confirmed pivot low pair as W2 / W4 proxies
    w4_low_price  = float(lows.iloc[-1]["PivotPrice"])
    w2_low_price  = float(lows.iloc[-2]["PivotPrice"]) if len(lows) >= 2 else w4_low_price
    w3_high_price = float(highs.iloc[-1]["PivotPrice"])

    w4_low_bar = int(lows.iloc[-1]["PivotBar"])
    w2_low_bar = int(lows.iloc[-2]["PivotBar"]) if len(lows) >= 2 else w4_low_bar - 10

    # Slope of base channel
    bar_diff = w4_low_bar - w2_low_bar
    if bar_diff == 0:
        return {"Channel_Base": "", "Channel_Top": "", "W5_Channel_Target": ""}

    slope = (w4_low_price - w2_low_price) / bar_diff

    # Project forward ~5-15 bars for W5 channel target
    bars_forward = max(5, bar_diff // 2)
    n_total = len(df)
    w5_bar = n_total + bars_forward

    # Base channel at W5 bar
    channel_base_w5 = w4_low_price + slope * bars_forward

    # Height = W3 top minus W2 low (channel width)
    channel_width = w3_high_price - w2_low_price
    channel_top_w5 = channel_base_w5 + channel_width

    return {
        "Channel_Base": round(channel_base_w5, 2),
        "Channel_Top":  round(channel_top_w5,  2),
        "W5_Channel_Target": round(channel_top_w5, 2),
    }


# ────────────────────────────────────────────────────────────────
# A5. FIBONACCI TIME PROJECTIONS (Frost Ch 4)
#     Projects WHEN the next wave might end
#     Uses Fib ratios: 0.382, 0.618, 1.0, 1.618 of prior wave duration
# ────────────────────────────────────────────────────────────────
def fibonacci_time_projection(pivots_df, df, horizon="long"):
    """
    Projects the end date/bar of the CURRENT wave using Fibonacci time ratios
    applied to the immediately prior wave's duration.
    Returns estimated bars_remaining and calendar date.
    """
    result = {
        "Wave_Duration_Bars":    "",
        "Wave_End_Est_Bars":     "",
        "Wave_End_Est_Date":     "",
        "Wave_End_Est_Price":    "",
        "Next_Wave_Start_Est":   "",
        "Fib_Time_Ratio_Used":   "",
    }

    if pivots_df is None or pivots_df.empty or df is None or len(df) < 20:
        return result

    all_pivots = pivots_df.sort_values("ConfirmBar")
    if len(all_pivots) < 3:
        return result

    # Get last 3 confirmed pivots to measure wave durations
    p3 = all_pivots.iloc[-1]
    p2 = all_pivots.iloc[-2]
    p1 = all_pivots.iloc[-3]

    prior_wave_bars  = int(p2["ConfirmBar"]) - int(p1["ConfirmBar"])
    current_wave_bars = int(len(df) - 1) - int(p2["ConfirmBar"])

    if prior_wave_bars <= 0:
        return result

    # Fibonacci time ratios from Frost & Prechter Ch 4
    fib_ratios = [0.382, 0.618, 1.0, 1.618, 2.618]
    best_ratio = None
    bars_remaining = None

    for ratio in fib_ratios:
        projected_total = prior_wave_bars * ratio
        remaining = projected_total - current_wave_bars
        if remaining > 0:
            best_ratio = ratio
            bars_remaining = int(remaining)
            break

    if bars_remaining is None or bars_remaining > 200:
        bars_remaining = prior_wave_bars  # fallback: same duration as prior wave

    # Convert bars to calendar date
    try:
        last_date = pd.to_datetime(df.index[-1])
        if horizon == "long":
            end_date = last_date + pd.tseries.offsets.Week(bars_remaining)
            next_wave_date = last_date + pd.tseries.offsets.Week(bars_remaining + 2)
        else:
            end_date = last_date + pd.tseries.offsets.BDay(bars_remaining)
            next_wave_date = last_date + pd.tseries.offsets.BDay(bars_remaining + 3)

        # Estimate price at wave end using linear channel slope
        close_arr = df["Close"].values
        slope_30 = (close_arr[-1] - close_arr[-min(30, len(close_arr))]) / min(30, len(close_arr))
        est_price = close_arr[-1] + slope_30 * bars_remaining * 0.5  # damped projection

        result["Wave_Duration_Bars"]  = current_wave_bars
        result["Wave_End_Est_Bars"]   = bars_remaining
        result["Wave_End_Est_Date"]   = str(end_date.date())
        result["Wave_End_Est_Price"]  = round(float(est_price), 2)
        result["Next_Wave_Start_Est"] = str(next_wave_date.date())
        result["Fib_Time_Ratio_Used"] = str(best_ratio) if best_ratio else "1.0"

    except Exception:
        pass

    return result


# ────────────────────────────────────────────────────────────────
# A6. WALK-FORWARD BACKTEST ENGINE
#     Slides a window back in time, runs ew_screen(), records
#     signal → outcome (did price go up 5%+ in next N bars?)
# ────────────────────────────────────────────────────────────────
def backtest_ticker(ticker, horizon="long", lookback_windows=8, fwd_bars=20):
    """
    Walk-forward backtest for a single ticker.
    Returns DataFrame of: Date | State | Signal | Fwd_Return | Hit
    """
    days   = 365 * 5 if horizon == "long" else 365
    intv   = "1wk"   if horizon == "long" else "1d"
    df_all = download_data(ticker, days, interval=intv)

    if df_all is None or len(df_all) < 100:
        return pd.DataFrame()

    rows = []
    step = max(5, len(df_all) // (lookback_windows + 1))

    for i in range(lookback_windows):
        end_idx = len(df_all) - fwd_bars - i * step
        if end_idx < 60:
            break

        df_slice = df_all.iloc[:end_idx].copy()
        result   = ew_screen(df_slice, horizon)

        if result is None:
            continue

        state  = result["Elliott_State"]
        action = result["Recommended_Action"]
        price_at_signal = df_all["Close"].iloc[end_idx - 1]
        price_fwd = df_all["Close"].iloc[min(end_idx + fwd_bars - 1, len(df_all) - 1)]
        fwd_ret   = (price_fwd - price_at_signal) / price_at_signal * 100

        is_buy = any(k in state.upper() for k in ["WAVE 2", "WAVE 3", "ACCUMULATION", "CAPITULATION", "BREAKOUT"])
        hit    = (fwd_ret > 5.0) if is_buy else (fwd_ret < -2.0)

        rows.append({
            "Date":        str(df_all.index[end_idx - 1].date()),
            "Elliott_State": state,
            "Signal":      action,
            "Price":       round(float(price_at_signal), 2),
            "Fwd_Return_%": round(fwd_ret, 2),
            "Hit":         "✅" if hit else "❌",
            "Is_Buy_Signal": is_buy,
        })

    return pd.DataFrame(rows)


def run_backtest_summary(tickers, horizon="long", lookback_windows=6, fwd_bars=20, max_workers=8):
    """
    Runs backtest across all tickers and returns summary DataFrame.
    Parallelized with a thread pool since each ticker's backtest is an
    independent, network I/O-bound operation (yfinance download).
    Per-ticker backtest logic is unchanged from the original.
    """
    all_rows = []

    def _run_one(ticker):
        try:
            bt = backtest_ticker(ticker, horizon, lookback_windows, fwd_bars)
            if not bt.empty:
                bt.insert(0, "Ticker", ticker)
                return bt
        except Exception:
            pass
        return None

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(_run_one, ticker) for ticker in tickers]
        for future in as_completed(futures):
            bt = future.result()
            if bt is not None:
                all_rows.append(bt)

    if not all_rows:
        return pd.DataFrame()

    combined = pd.concat(all_rows, ignore_index=True)
    return combined



def ew_state_meta(state, df, horizon):
    """Return the 6 new EW state columns for a given state + dataframe."""
    key = str(state).strip().upper()
    dur_bars = _current_state_bars(df, key, horizon)
    dur_min, dur_max = _STATE_DURATION_BARS.get(key, (max(dur_bars, 1), max(dur_bars, 1)))
    dur_human = _bars_to_human(dur_min, dur_max, horizon) if dur_bars > 0 else "unknown"
    next_info = _NEXT_STATE_MAP.get(key)
    if next_info:
        next_state, eta_min, eta_max = next_info
        eta_human = _bars_to_human(eta_min, eta_max, horizon)
        eta_bars_display = f"{eta_min}-{eta_max}"
    else:
        next_state, eta_bars_display, eta_human = "—", "—", "—"
    trigger = _TRANSITION_TRIGGER_MAP.get(key, "Watch RSI + MACD + price vs EMAs")
    return {
        "State_Duration_Bars":  dur_bars if dur_bars > 0 else "—",
        "State_Duration_Human": dur_human,
        "Next_EW_State":        next_state,
        "Next_State_ETA_Bars":  eta_bars_display,
        "Next_State_ETA_Human": eta_human,
        "Transition_Trigger":   trigger,
    }

def ew_screen(df, horizon="long"):
    if df is None or len(df) < 50:
        return None

    c = df["Close"].values
    h = df["High"].values
    l = df["Low"].values
    v = df["Volume"].values
    n = len(c)

    p200 = min(200, max(20, n // 2))
    p50 = min(50, max(10, n // 3))
    p20 = 20 if n >= 20 else max(5, n // 4)

    m200 = ema(c, p200)
    m50 = ema(c, p50)
    m20 = ema(c, p20)

    above_200 = c[-1] > m200[-1]
    above_50 = c[-1] > m50[-1]
    above_20 = c[-1] > m20[-1]

    bull_ma_score = int(above_200) + int(above_50) + int(above_20)
    ma_bull_align = m200[-1] < m50[-1] < m20[-1] < c[-1]
    ma_bear_align = m200[-1] > m50[-1] > m20[-1] > c[-1]

    rsi_arr = rsi_calc(c)
    if len(rsi_arr) == 0:
        return None

    rsi_cur = rsi_arr[-1]
    rsi_prev = rsi_arr[-5] if len(rsi_arr) > 5 else rsi_cur
    rsi_rising = rsi_cur > rsi_prev

    macd_line, macd_signal, macd_hist = macd_calc(c)
    macd_bull = macd_line[-1] > macd_signal[-1] and macd_line[-1] > 0
    macd_bear = macd_line[-1] < macd_signal[-1] and macd_line[-1] < 0
    macd_expanding = len(macd_hist) > 3 and macd_hist[-1] > macd_hist[-2] > macd_hist[-3]
    macd_shrinking = len(macd_hist) > 3 and macd_hist[-1] < macd_hist[-2] < macd_hist[-3]
    macd_div_bull = macd_line[-1] > macd_signal[-1] and macd_line[-1] < 0
    macd_div_bear = macd_line[-1] < macd_signal[-1] and macd_line[-1] > 0
    macd_cross_up = len(macd_line) > 2 and macd_line[-1] > macd_signal[-1] and macd_line[-2] <= macd_signal[-2]

    vol_ma20 = pd.Series(v).rolling(20).mean().values
    vol_ratio = v[-1] / (vol_ma20[-1] + 1e-6) if len(vol_ma20) else 0
    vol_rising = np.mean(v[-5:]) > np.mean(v[-20:]) if len(v) >= 20 else False
    vol_dry = vol_ratio < 0.75
    vol_surge = vol_ratio > 1.1   # Wave 3 = sustained volume, not just spikes
    vol_climax = vol_ratio > 2.5

    obv = obv_calc(c, v)
    obv_ma = pd.Series(obv).rolling(20).mean().values
    obv_rising = obv[-1] > obv_ma[-1] and obv[-1] > obv[-5] if len(obv) > 20 and len(obv_ma) > 0 else False
    obv_bull_div = c[-1] < c[-10] and obv[-1] > obv[-10] if len(c) > 10 else False
    obv_bear_div = c[-1] > c[-10] and obv[-1] < obv[-10] if len(c) > 10 else False

    lb = min(260, n) if horizon == "long" else min(100, n)  # long=~5yr weekly, mid/short=100 daily
    swg_hi = np.max(h[-lb:])
    swg_lo = np.min(l[-lb:])
    fib_range = swg_hi - swg_lo if swg_hi > swg_lo else 0.0001

    f382 = swg_hi - fib_range * 0.382
    f500 = swg_hi - fib_range * 0.500
    f618 = swg_hi - fib_range * 0.618
    f786 = swg_hi - fib_range * 0.786
    f100 = swg_lo
    f161 = swg_hi + fib_range * 0.618

    in_w2_zone = False  # defined after rule1_ok below
    near_bottom = c[-1] <= f100 * 1.05 and not above_50

    hi52 = np.max(h[-252:]) if n >= 252 else np.max(h)
    lo52 = np.min(l[-252:]) if n >= 252 else np.min(l)
    pct_from_hi = (hi52 - c[-1]) / hi52 * 100 if hi52 != 0 else 0
    pct_from_lo = (c[-1] - lo52) / lo52 * 100 if lo52 != 0 else 0

    range10 = (np.max(h[-10:]) - np.min(l[-10:])) if n >= 10 else 0
    range20 = (np.max(h[-20:]) - np.min(l[-20:])) if n >= 20 else 0
    range40 = (np.max(h[-40:]) - np.min(l[-40:])) if n >= 40 else 0
    range_contracting = range20 > 0 and range40 > 0 and range10 < range20 * 0.7 and range20 < range40 * 0.7

    selling_climax = vol_climax and not above_50 and rsi_cur < 30
    blowoff = vol_climax and above_200 and rsi_cur > 75
    wave3_active = macd_bull and macd_expanding and bull_ma_score >= 2 and (vol_surge or vol_ratio > 0.9)
    wave5_warn = (
        c[-1] >= np.max(c[-20:]) * 0.995 and
        vol_dry and
        (rsi_cur > 65 or obv_bear_div) and
        above_200
        # macd_div_bear removed: Wave 5 top warning must fire BEFORE MACD rolls over
    )
    b_trap = ((not above_200) and len(c) > 5 and c[-1] > c[-5] and macd_shrinking and vol_ratio < 1.1)
    accumulation = above_50 and not above_200 and 35 < rsi_cur < 58 and (obv_rising or rsi_rising)  # EW-V3 fix: early Dow Phase1 OBV may lag

    rule1_ok = c[-1] > f100
    in_w2_zone = (f618 * 0.97 <= c[-1] <= f618 * 1.03) and rule1_ok  # EW-V2 fix: rule1_ok guards Wave2, not above_200
    w4_overlap_risk = above_200 and c[-1] < f786 and 35 < rsi_cur < 50

    pivot_info = recent_pivot_summary(
        df,
        left_bars=PIVOT_LEFT_BARS,
        right_bars=PIVOT_RIGHT_BARS,
        near_bars=PIVOT_NEAR_BARS
    )

    score = 0
    signals = []

    if horizon == "long":
        if selling_climax:
            score += 20; signals.append("Selling Climax")
        if near_bottom:
            score += 15; signals.append("Near Cycle Bottom")
        if obv_bull_div:
            score += 15; signals.append("OBV Bullish Divergence")
        if macd_div_bull:
            score += 10; signals.append("MACD Below-Zero Bull Cross")
        if ma_bull_align:
            score += 10; signals.append("All MAs Bull Aligned")
        if bull_ma_score >= 2:
            score += 10; signals.append(f"Dow Score {bull_ma_score}/3")
        if rsi_cur < 40 and rsi_rising:
            score += 10; signals.append("RSI Recovering from Oversold")
        if pct_from_lo < 25:
            score += 5; signals.append("Near 52-Week Low")
        if above_200:
            score += 5; signals.append("Above 200 EMA")

        if selling_climax and obv_bull_div:
            state = "CAPITULATION — Bull Cycle Starting"
        elif accumulation or (near_bottom and obv_rising):
            state = "ACCUMULATION — Dow Phase 1"
        elif (ma_bull_align or bull_ma_score == 3) and macd_bull and macd_expanding:
            state = "WAVE 3 — Strongest Wave (Ride It)"  # EW-V4 fix: catch early W3 when 50EMA still below 200EMA
        elif bull_ma_score >= 2 and (macd_div_bull or in_w2_zone):
            state = "WAVE 2 Buy Zone — Best Long Entry"
        elif wave5_warn:
            state = "WAVE 5 — Late Bull, Start Trimming"
        elif blowoff:
            state = "BLOWOFF TOP — Exit All Longs"
        elif ma_bear_align or (not above_200 and not above_50 and rsi_cur < 45):
            state = "BEAR MARKET — Avoid / Cash"  # EW-V5 fix: partial bear (dead-cat bounce won't fool it)
        elif b_trap:
            state = "B-WAVE BULL TRAP — Fake Rally"
        else:
            state = "TRANSITIONING / Sideways"

        action = (
            "BUY / BUILD POSITION" if any(x in state for x in ["CAPITULATION", "ACCUMULATION", "WAVE 2"])
            else "HOLD / ADD DIPS" if "WAVE 3" in state
            else "TRIM / EXIT" if any(x in state for x in ["WAVE 5", "BLOWOFF", "BEAR", "TRAP"])
            else "WATCH"
        )
        hold_period = "1-5 Years"
        risk = f"Stop: 200 EMA ({fmt_price(m200[-1])}) | Wave2 invalidation: {fmt_price(f100)}"

    elif horizon == "mid":
        if in_w2_zone:
            score += 20; signals.append("In Wave 2 61.8% Zone")
        if obv_bull_div:
            score += 15; signals.append("OBV Bullish Divergence")
        if macd_div_bull or macd_cross_up:
            score += 15; signals.append("MACD Bullish Cross")
        if bull_ma_score >= 2:
            score += 10; signals.append(f"Dow Score {bull_ma_score}/3")
        if 40 <= rsi_cur <= 60 and rsi_rising:
            score += 10; signals.append("RSI Neutral Rising")
        if vol_rising:
            score += 10; signals.append("Volume Expanding")
        if above_50 and not above_200:
            score += 10; signals.append("Reclaiming 50 EMA")
        if range_contracting:
            score += 5; signals.append("Triangle Compression")
        if pct_from_hi < 30 and above_200:
            score += 5; signals.append("Near High in Bull Trend")

        if in_w2_zone and obv_bull_div:
            state = "WAVE 2 Buy Zone (Best Risk:Reward)"
        elif macd_cross_up and above_50 and vol_rising:
            state = "WAVE 3 Starting — Enter Now"
        elif wave3_active:
            state = "WAVE 3 Active — Hold / Trail"
        elif range_contracting and above_50:
            state = "TRIANGLE — Wait for Thrust"
        elif macd_bull and ma_bull_align:
            state = "WAVE 3 Active"
        elif wave5_warn:
            state = "WAVE 5 — Reduce Risk"
        elif b_trap:
            state = "B-WAVE TRAP — Stand Aside"
        elif accumulation:
            state = "ACCUMULATION — Build Slowly"
        else:
            state = "WATCH LIST — Setup Forming"

        action = (
            "BUY HERE" if "WAVE 2" in state
            else "ENTER LONG" if "Starting" in state
            else "HOLD" if "WAVE 3 Active" in state or "Hold" in state
            else "BUY ON BREAKOUT" if "TRIANGLE" in state
            else "REDUCE / EXIT" if any(x in state for x in ["WAVE 5", "TRAP"])
            else "WATCH"
        )
        hold_period = "2 Weeks - 4 Months"
        risk = f"Stop: 61.8% Fib ({fmt_price(f618)}) | 50 EMA ({fmt_price(m50[-1])})"

    else:
        if macd_bull and macd_expanding:
            score += 25; signals.append("MACD Bull + Expanding")
        if bull_ma_score == 3:
            score += 20; signals.append("All 3 MAs Bullish")
        if vol_ratio > 1.5:
            score += 15; signals.append(f"Volume Surge {round(vol_ratio,1)}x")
        if 55 <= rsi_cur <= 75:
            score += 10; signals.append("RSI Bullish Momentum")
        if obv_rising:
            score += 10; signals.append("OBV Rising")
        if macd_cross_up:
            score += 10; signals.append("Fresh MACD Bull Cross")
        if above_20 and above_50:
            score += 5; signals.append("Price Above 20 & 50 EMA")
        if vol_ratio > 2.0:
            score += 5; signals.append("Volume Climax Confirmation")

        if macd_bull and macd_expanding and bull_ma_score >= 2 and vol_surge:
            state = "3rd of 3rd Wave — DO NOT SHORT"
        elif macd_cross_up and vol_ratio > 1.3 and above_50:
            state = "WAVE 3 Breakout — Enter Now"
        elif range_contracting and macd_cross_up:
            state = "TRIANGLE THRUST — Buy Breakout"
        elif macd_bull and vol_rising and bull_ma_score >= 2:
            state = "WAVE 3 Active — Ride It"
        elif wave5_warn:
            state = "WAVE 5 TOP — Start Exiting"
        elif blowoff:
            state = "BLOWOFF — Exit Immediately"
        else:
            state = "SETUP FORMING — Wait for Signal"

        action = (
            "BUY NOW / LONG" if any(x in state for x in ["3rd of 3rd", "Breakout", "THRUST"])
            else "HOLD LONG" if "WAVE 3 Active" in state or "Ride It" in state
            else "EXIT / SHORT" if any(x in state for x in ["WAVE 5", "BLOWOFF"])
            else "WAIT"
        )
        hold_period = "Hours - 2 Weeks"
        risk = f"Trail: 20 EMA ({fmt_price(m20[-1])}) | Stop: 50 EMA ({fmt_price(m50[-1])})"

    explain = wave_guidance(
        state=state,
        close=c[-1],
        f382=f382,
        f500=f500,
        f618=f618,
        f786=f786,
        f100=f100,
        f161=f161,
        hi52=hi52,
        lo52=lo52,
        m20=m20[-1],
        m50=m50[-1],
        m200=m200[-1]
    )


    # ── Frost & Prechter: Channel + Fib Time (injected before return) ────
    _piv_df = detect_confirmed_pivots(df, PIVOT_LEFT_BARS, PIVOT_RIGHT_BARS)
    channel_info = build_wave_channel(_piv_df, df)
    fib_time     = fibonacci_time_projection(_piv_df, df, horizon)

    return {
        "EW_Score": min(score, 100),
        "Trade_Quality": explain["Trade_Quality"],
        "Elliott_State": state,
        "Current_Price": fmt_price(c[-1]),
        "Current_Wave_Target": explain["Current_Wave_Target"],
        "Ideal_Entry_Price": explain["Ideal_Entry_Price"],
        "Pullback_Buy_Zone": explain["Pullback_Buy_Zone"],
        "Stop_Loss": explain["Stop_Loss"],
        "Invalidation_Level": explain["Invalidation_Level"],
        "Extension_Target": explain["Extension_Target"],
        "If_Extended_Then": explain["If_Extended_Then"],
        "Beginner_Note": explain["Beginner_Note"],
        "RSI": round(float(rsi_cur), 1),
        "Vol_Ratio": round(float(vol_ratio), 2),
        "Dow_MA_Score": f"{bull_ma_score}/3",
        "OBV_Divergence": "YES" if obv_bull_div else "No",
        "MACD_Signal": "BULL" if macd_bull else "CROSS↑" if macd_cross_up else "BEAR" if macd_bear else "NEUTRAL",
        "Pct_from_52w_High": round(float(pct_from_hi), 1),
        "Pct_from_52w_Low": round(float(pct_from_lo), 1),
        "Key_Signals": ", ".join(signals) if signals else "None",
        "52w_High": fmt_price(hi52),
        "52w_Low": fmt_price(lo52),
        "Fib_61.8%": fmt_price(f618),
        "Fib_38.2%": fmt_price(f382),
        "Fib_161.8%_Target": fmt_price(f161),
        "Recommended_Action": action,
        "Hold_Period": hold_period,
        "Risk_Management": risk,
        "Rule1_Wave2_OK": "OK" if rule1_ok else "VIOLATED",
        "W4_Overlap_Risk": "YES" if w4_overlap_risk else "No",
        "Pivot_Low": pivot_info["Pivot_Low"],
        "Pivot_Low_Date": pivot_info["Pivot_Low_Date"],
        "Pivot_Low_Price": pivot_info["Pivot_Low_Price"],
        "Pivot_Low_Bars_Ago": pivot_info["Pivot_Low_Bars_Ago"],
        "Pivot_High": pivot_info["Pivot_High"],
        "Pivot_High_Date": pivot_info["Pivot_High_Date"],
        "Pivot_High_Price": pivot_info["Pivot_High_Price"],
        "Pivot_High_Bars_Ago": pivot_info["Pivot_High_Bars_Ago"],

        # ── NEW: Frost & Prechter Complete Rule Set ──────────────────────
        "Wave_Degree":          classify_wave_degree(horizon, n,
                                    "1wk" if horizon == "long" else "1d"),
        "Corrective_Pattern":   classify_corrective_pattern(
                                    detect_confirmed_pivots(df, PIVOT_LEFT_BARS, PIVOT_RIGHT_BARS),
                                    c, swg_hi, swg_lo),
        "Diagonal_Triangle":    detect_diagonal_triangle(df, horizon) or "None",
        "Channel_Base":         channel_info.get("Channel_Base", ""),
        "Channel_Top":          channel_info.get("Channel_Top", ""),
        "W5_Channel_Target":    channel_info.get("W5_Channel_Target", ""),
        "Wave_Duration_Bars":   fib_time.get("Wave_Duration_Bars", ""),
        "Wave_End_Est_Bars":    fib_time.get("Wave_End_Est_Bars", ""),
        "Wave_End_Est_Date":    fib_time.get("Wave_End_Est_Date", ""),
        "Wave_End_Est_Price":   fib_time.get("Wave_End_Est_Price", ""),
        "Next_Wave_Start_Est":  fib_time.get("Next_Wave_Start_Est", ""),
        "Fib_Time_Ratio":       fib_time.get("Fib_Time_Ratio_Used", ""),
                **ew_state_meta(state, df, horizon)
    }




# ════════════════════════════════════════════════════════════════
# U.S. EXCHANGE UNIVERSE BUILDERS
# ════════════════════════════════════════════════════════════════

# ── Public ETF Master List ───────────────────────────────────────
ETF_UNIVERSE = {
    # Broad Market
    "SPY": "S&P 500 ETF", "QQQ": "Nasdaq-100 ETF", "DIA": "Dow 30 ETF",
    "IWM": "Russell 2000 ETF", "VTI": "Total US Market ETF", "VT": "Total World ETF",
    # Sector SPDRs
    "XLK": "Technology", "XLF": "Financials", "XLV": "Health Care",
    "XLE": "Energy", "XLY": "Consumer Discret.", "XLP": "Consumer Staples",
    "XLI": "Industrials", "XLU": "Utilities", "XLB": "Materials",
    "XLRE": "Real Estate", "XLC": "Communication",
    # Thematic
    "SOXX": "Semiconductors", "SMH": "Semiconductors 2",
    "BOTZ": "Robotics & AI", "ROBO": "Robotics", "ARKK": "ARK Innovation",
    "ARKG": "ARK Genomics", "ARKW": "ARK Web 3.0",
    "BUG": "Cybersecurity", "CIBR": "Cybersecurity 2",
    "SKYY": "Cloud Computing", "WCLD": "Cloud 2", "FDN": "Internet",
    "GLD": "Gold", "SLV": "Silver", "USO": "Oil", "UNG": "Nat Gas",
    "TAN": "Solar", "FAN": "Wind", "ICLN": "Clean Energy", "URA": "Uranium",
    "LIT": "Lithium & Battery", "DBA": "Agriculture",
    "AGG": "Bonds Agg", "TLT": "20yr Treasury", "IEF": "7-10yr Treasury",
    "SHY": "1-3yr Treasury", "LQD": "IG Corp Bonds", "HYG": "High Yield",
    "BIL": "T-Bills", "BNDX": "Intl Bonds", "EMB": "EM Bonds",
    "ITA": "Defense & Aerospace", "PPA": "Defense 2", "DFEN": "Defense 3x",
    "FXI": "China Large Cap", "MCHI": "MSCI China",
    "EWG": "Germany", "EWJ": "Japan", "EWU": "UK", "INDA": "India",
    "EWW": "Mexico", "EWA": "Australia", "EWC": "Canada",
    "VNQ": "REIT", "IYR": "Real Estate 2",
    "GDX": "Gold Miners", "GDXJ": "Jr Gold Miners",
    "XOP": "Oil & Gas E&P", "OIH": "Oil Services",
    "IBB": "Biotech", "XBI": "Biotech 2",
    "KRE": "Regional Banks", "KBE": "Banks",
    "JETS": "Airlines", "AWAY": "Travel",
    "ARKX": "Space Exploration", "UFO": "Space",
}

def get_etf_universe():
    rows = [{"Ticker": k, "Company": v, "Sector": "ETF", "SubIndustry": "Public ETF"}
            for k, v in ETF_UNIVERSE.items()]
    return pd.DataFrame(rows)


def get_nyse_listed():
    """Fetch NYSE-listed common stocks from NASDAQ trader file."""
    url = "https://www.nasdaqtrader.com/dynamic/SymDir/otherlisted.txt"
    try:
        r = requests.get(url, headers=HEADERS, timeout=30)
        r.raise_for_status()
        lines = r.text.splitlines()
        rows = [line.split("|") for line in lines if "|" in line]
        if len(rows) < 2:
            return pd.DataFrame()
        df = pd.DataFrame(rows[1:], columns=rows[0])
        df = df[df.get("ACT Symbol", pd.Series(dtype=str)).astype(str) != "File Creation Time"]
        df = df[df.get("ETF", "N") == "N"] if "ETF" in df.columns else df
        # Keep only NYSE and NYSE ARCA
        if "Exchange" in df.columns:
            df = df[df["Exchange"].isin(["N", "A", "P"])]
        sym_col = "ACT Symbol" if "ACT Symbol" in df.columns else "Symbol"
        name_col = "Security Name" if "Security Name" in df.columns else "Company Name"
        if sym_col not in df.columns:
            return pd.DataFrame()
        df["Ticker"] = normalize_ticker_series(df[sym_col])
        df["Company"] = df[name_col] if name_col in df.columns else ""
        df["Sector"] = ""
        df["SubIndustry"] = ""
        return df[["Ticker", "Company", "Sector", "SubIndustry"]].dropna(subset=["Ticker"]).reset_index(drop=True)
    except Exception as e:
        print(f"  [WARN] NYSE fetch failed: {e}")
        return pd.DataFrame()


def build_full_us_universe():
    """
    Builds 4 exchange universes + ETF universe.
    Returns dict: {exchange_code -> DataFrame}
    """
    universes = {}

    print("  Fetching S&P 500...")
    try:
        universes["SP500"] = get_sp500_constituents()
        print(f"    SP500: {len(universes['SP500'])} tickers")
    except Exception as e:
        print(f"    SP500 failed: {e}")
        universes["SP500"] = pd.DataFrame()

    print("  Fetching NASDAQ (broad)...")
    try:
        nasdaq_raw = get_nasdaq_listed_common()
        universes["NASDAQ"] = nasdaq_raw
        print(f"    NASDAQ: {len(universes['NASDAQ'])} tickers")
    except Exception as e:
        print(f"    NASDAQ failed: {e}")
        universes["NASDAQ"] = pd.DataFrame()

    print("  Fetching Russell 2000 proxy (S&P 600)...")
    try:
        universes["RUSSELL"] = get_sp600_constituents()
        print(f"    RUSSELL: {len(universes['RUSSELL'])} tickers")
    except Exception as e:
        print(f"    RUSSELL failed: {e}")
        universes["RUSSELL"] = pd.DataFrame()

    print("  Fetching Dow 30 + NYSE...")
    try:
        dow = get_dow30_constituents()
        nyse = get_nyse_listed()
        universes["DOW_NYSE"] = pd.concat([dow, nyse], ignore_index=True).drop_duplicates(subset=["Ticker"])
        print(f"    DOW_NYSE: {len(universes['DOW_NYSE'])} tickers")
    except Exception as e:
        print(f"    DOW_NYSE failed: {e}")
        universes["DOW_NYSE"] = pd.DataFrame()

    print("  Building ETF universe...")
    universes["ETF"] = get_etf_universe()
    print(f"    ETF: {len(universes['ETF'])} tickers")

    return universes


def _scan_single_ticker(tkr, universe_df):
    """Run all timeframe downloads + EW screens for a single ticker.
    Same logic as the original sequential loop body, extracted so it
    can be executed concurrently (I/O-bound network calls)."""
    meta_m = universe_df.loc[universe_df["Ticker"] == tkr]
    if meta_m.empty:
        cpy, sec, sub = "", "", ""
    else:
        row = meta_m.iloc[0]
        cpy  = row.get("Company", "")
        sec  = row.get("Sector", "")
        sub  = row.get("SubIndustry", "")

    rl = rm = rs = rmn = None
    try:
        fund_status = get_fundamental_status_with_history(tkr)

        dfl  = download_data(tkr, 365 * 5, interval="1wk")
        rl   = ew_screen(dfl, "long")
        if rl:
            rl.update({"Ticker": tkr, "Company": cpy, "Sector": sec, "SubIndustry": sub, "Fundamentally strong": fund_status})

        dfm  = download_data(tkr, 365, interval="1d")
        rm   = ew_screen(dfm, "mid")
        if rm:
            rm.update({"Ticker": tkr, "Company": cpy, "Sector": sec, "SubIndustry": sub, "Fundamentally strong": fund_status})

        dfs  = download_data(tkr, 90, interval="1d")
        rs   = ew_screen(dfs, "short")
        if rs:
            rs.update({"Ticker": tkr, "Company": cpy, "Sector": sec, "SubIndustry": sub, "Fundamentally strong": fund_status})

        dfmn = download_data_monthly(tkr, years=10)
        rmn  = ew_screen(dfmn, "long")
        if rmn:
            rmn.update({"Ticker": tkr, "Company": cpy, "Sector": sec, "SubIndustry": sub, "Timeframe": "Monthly", "Fundamentally strong": fund_status})
    except Exception:
        pass

    return tkr, rl, rm, rs, rmn


def run_scan_for_universe(universe_df, universe_label, max_workers=8):
    """Run the full EW scan for a given universe DataFrame.
    Uses a thread pool to parallelize the network-bound per-ticker
    downloads (yfinance/requests), since these calls release the GIL
    while waiting on I/O. All per-ticker computation logic (download
    windows, EW screening, field mapping) is unchanged from the
    original sequential version."""
    tickers_list = universe_df["Ticker"].dropna().unique().tolist()

    r_lt_u, r_mt_u, r_st_u, r_mn_u = [], [], [], []

    print(f"\n{'=' * 72}")
    print(f"SCANNING {universe_label.upper()} ({len(tickers_list)} tickers)")
    print(f"{'=' * 72}")

    done_count = 0
    total = len(tickers_list)

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {executor.submit(_scan_single_ticker, tkr, universe_df): tkr for tkr in tickers_list}
        for future in as_completed(futures):
            tkr = futures[future]
            done_count += 1
            print(f"  [{done_count:04d}/{total}] {tkr}...", end="\r")
            try:
                _, rl, rm, rs, rmn = future.result()
            except Exception:
                continue
            if rl:
                r_lt_u.append(rl)
            if rm:
                r_mt_u.append(rm)
            if rs:
                r_st_u.append(rs)
            if rmn:
                r_mn_u.append(rmn)

    print(f"  Done: {len(r_lt_u)} LT | {len(r_mt_u)} MT | {len(r_st_u)} ST | {len(r_mn_u)} Monthly")
    return r_lt_u, r_mt_u, r_st_u, r_mn_u


# ════════════════════════════════════════════════════════════════

# ────────────────────────────────────────────────────────────────
# USER INPUTS — CHANGE ONLY SCRIPT_DIR
# ────────────────────────────────────────────────────────────────

CUSTOM_UNIVERSE_NAME = "NASDAQ Composite"
CUSTOM_UNIVERSE_CODE = "NASDAQ_Composite"

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILENAME = os.path.join(SCRIPT_DIR, "Elliott_Wave_NASDAQ_Composite_Master_Workbook.xlsx")

# Pivot settings
PIVOT_LEFT_BARS  = 5
PIVOT_RIGHT_BARS = 5
PIVOT_NEAR_BARS  = 2

HEADERS = {
    "User-Agent": "ElliottWaveScreener/1.0 (educational use)"
}

# Suppress yfinance & pandas warnings
CUSTOM_TICKERS = []  # not used in this script; kept for compatibility

# ────────────────────────────────────────────────────────────────
# EXECUTION BLOCK — NASDAQ Composite UNIVERSE
# ────────────────────────────────────────────────────────────────

UNIVERSE_NAME = CUSTOM_UNIVERSE_NAME
UNIVERSE_CODE = CUSTOM_UNIVERSE_CODE

print("=" * 72)
print(f"FETCHING NASDAQ Composite CONSTITUENTS...")
print("=" * 72)

try:
    universe_df = get_nasdaq_composite_proxy()
    print(f"  Loaded {len(universe_df)} tickers from NASDAQ Composite")
except Exception as _fetch_err:
    print(f"  [ERROR] Could not fetch NASDAQ Composite constituents: {_fetch_err}")
    universe_df = pd.DataFrame(columns=["Ticker","Company","Sector","SubIndustry"])

print("=" * 72)
print(f"SCANNING {UNIVERSE_NAME} ({len(universe_df)} tickers)")
print("=" * 72)

results_lt, results_mt, results_st, results_mn = run_scan_for_universe(
    universe_df, UNIVERSE_NAME, max_workers=16
)

TICKERS = universe_df["Ticker"].tolist()

print("\nScan complete.")
print(f"  Long-Term results  : {len(results_lt)}")
print(f"  Mid-Term  results  : {len(results_mt)}")
print(f"  Short-Term results : {len(results_st)}")
print(f"  Monthly results    : {len(results_mn)}")



# ────────────────────────────────────────────────────────────────
# BUILD DATAFRAMES
# ────────────────────────────────────────────────────────────────
col_order = [
    "Ticker","Company","Sector","SubIndustry","Fundamentally strong",
    "EW_Score","Trade_Quality","Elliott_State","Current_Price",
    "Current_Wave_Target","Ideal_Entry_Price","Pullback_Buy_Zone",
    "Stop_Loss","Invalidation_Level","Extension_Target","If_Extended_Then",
    "Beginner_Note","RSI","Vol_Ratio","Dow_MA_Score","OBV_Divergence","MACD_Signal",
    "Pct_from_52w_High","Pct_from_52w_Low","Key_Signals",
    "52w_High","52w_Low","Fib_61.8%","Fib_38.2%","Fib_161.8%_Target",
    "Recommended_Action","Hold_Period","Risk_Management",
    "Rule1_Wave2_OK","W4_Overlap_Risk",
    "Pivot_Low","Pivot_Low_Date","Pivot_Low_Price","Pivot_Low_Bars_Ago",
    "Pivot_High","Pivot_High_Date","Pivot_High_Price","Pivot_High_Bars_Ago",
    "State_Duration_Bars","State_Duration_Human",
    "Next_EW_State","Next_State_ETA_Bars","Next_State_ETA_Human",
    "Transition_Trigger",
    "Wave_Degree",
    "Corrective_Pattern",
    "Diagonal_Triangle",
    "Channel_Base",
    "Channel_Top",
    "W5_Channel_Target",
    "Wave_Duration_Bars",
    "Wave_End_Est_Bars",
    "Wave_End_Est_Date",
    "Wave_End_Est_Price",
    "Next_Wave_Start_Est",
    "Fib_Time_Ratio"
]


# ════════════════════════════════════════════════════════════════
# MULTI-TIMEFRAME ALIGNMENT — Helper + Builder
# ════════════════════════════════════════════════════════════════

def multi_tf_alignment_score(mn_state, lt_state, mt_state, st_state):
    mn = str(mn_state).upper()
    lt = str(lt_state).upper()
    mt = str(mt_state).upper()
    st = str(st_state).upper()

    # Hard conflict: monthly or weekly is bearish
    if any(x in mn for x in ["BEAR MARKET", "BLOWOFF", "B-WAVE TRAP"]):
        return "MONTHLY CONFLICT - SKIP"
    if any(x in lt for x in ["BEAR MARKET", "BLOWOFF", "B-WAVE TRAP"]):
        return "HIGHER TF CONFLICT - SKIP"

    # PLATINUM: All 4 aligned at bottom
    if (any(x in mn for x in ["CAPITULATION", "ACCUMULATION"]) and
        any(x in lt for x in ["CAPITULATION", "ACCUMULATION", "WAVE 2"]) and
        any(x in mt for x in ["ACCUMULATION", "WATCH LIST", "WAVE 2"]) and
        any(x in st for x in ["SETUP FORMING", "WAVE 3 BREAKOUT", "WAVE 2", "BREAKOUT", "3RD OF 3RD"])):
        return "PLATINUM - Monthly+Weekly+Daily Bottom Aligned"

    # GOLD: Monthly+Weekly aligned, Daily confirming
    if (any(x in mn for x in ["WAVE 2", "WAVE 3", "ACCUMULATION"]) and
        any(x in lt for x in ["CAPITULATION", "ACCUMULATION", "WAVE 2", "WAVE 3"]) and
        any(x in mt for x in ["WAVE 2", "WAVE 3 STARTING", "WATCH LIST", "WAVE 3 ACTIVE"]) and
        any(x in st for x in ["WAVE 3 BREAKOUT", "SETUP FORMING", "WAVE 3 ACTIVE", "RIDE IT", "3RD OF 3RD"])):
        return "GOLD - Monthly+Weekly Aligned, Daily Confirming"

    # SILVER: Monthly+Daily aligned (Weekly may lag)
    if (any(x in mn for x in ["WAVE 2", "WAVE 3", "ACCUMULATION"]) and
        any(x in mt for x in ["WAVE 2", "WAVE 3 STARTING", "WAVE 3 ACTIVE"]) and
        any(x in st for x in ["WAVE 3 BREAKOUT", "3RD OF 3RD", "RIDE IT"])):
        return "SILVER - Monthly+Daily Aligned"

    # STRONG: Weekly+Daily aligned (no monthly conflict)
    if (any(x in lt for x in ["WAVE 2", "WAVE 3", "ACCUMULATION"]) and
        any(x in mt for x in ["WAVE 2", "WAVE 3 STARTING", "WATCH LIST", "WAVE 3 ACTIVE"]) and
        any(x in st for x in ["WAVE 3 BREAKOUT", "SETUP FORMING", "WAVE 3 ACTIVE", "RIDE IT", "3RD OF 3RD"])):
        return "STRONG - Weekly+Daily Aligned, No Conflict"

    if (any(x in lt for x in ["WAVE 3"]) and
        any(x in mt for x in ["WAVE 3"]) and
        any(x in st for x in ["WAVE 3", "BREAKOUT", "RIDE", "3RD OF 3RD"])):
        return "GOOD - All 3 TF in Wave 3"

    if any(x in st for x in ["WAVE 5", "BLOWOFF"]):
        return "LATE - Short-term extended, reduce size"

    return "NEUTRAL - Wait for better alignment"


def build_multi_tf_sheet(df_monthly, df_long, df_mid, df_short):
    dfs = [df_monthly, df_long, df_mid, df_short]
    if any(d is None or d.empty for d in dfs):
        # Try without monthly if not available
        if df_long is None or df_long.empty or df_mid is None or df_mid.empty or df_short is None or df_short.empty:
            return pd.DataFrame()

    # Base from df_long
    base_cols_lt = [c for c in [
        "Ticker", "Company", "Sector", "SubIndustry",
        "Elliott_State", "EW_Score", "Trade_Quality", "Current_Price",
        "Ideal_Entry_Price", "Stop_Loss", "Extension_Target",
        "Recommended_Action", "Hold_Period", "Risk_Management"
    ] if c in df_long.columns]

    merged = df_long[base_cols_lt].copy()
    merged = merged.rename(columns={
        "Elliott_State":      "LT_State",
        "EW_Score":           "LT_Score",
        "Trade_Quality":      "LT_Quality",
        "Ideal_Entry_Price":  "LT_Entry",
        "Stop_Loss":          "LT_Stop",
        "Recommended_Action": "LT_Action",
        "Hold_Period":        "LT_Hold",
        "Risk_Management":    "LT_Risk",
    })

    # Monthly
    has_monthly = df_monthly is not None and not df_monthly.empty
    if has_monthly:
        mn_pick = [c for c in ["Ticker", "Elliott_State", "EW_Score", "Trade_Quality"] if c in df_monthly.columns]
        mn_cols = df_monthly[mn_pick].copy()
        mn_cols = mn_cols.rename(columns={
            "Elliott_State": "MN_State",
            "EW_Score":      "MN_Score",
            "Trade_Quality": "MN_Quality",
        })
        merged = merged.merge(mn_cols, on="Ticker", how="left")
        merged["MN_State"]   = merged.get("MN_State",   "N/A").fillna("N/A")
        merged["MN_Score"]   = merged.get("MN_Score",   0).fillna(0)
        merged["MN_Quality"] = merged.get("MN_Quality", "N/A").fillna("N/A")
    else:
        merged["MN_State"]   = "N/A"
        merged["MN_Score"]   = 0
        merged["MN_Quality"] = "N/A"

    mt_pick = [c for c in [
        "Ticker", "Elliott_State", "EW_Score", "Trade_Quality",
        "Ideal_Entry_Price", "Stop_Loss", "Recommended_Action", "Hold_Period"
    ] if c in df_mid.columns]
    mt_cols = df_mid[mt_pick].copy()
    mt_cols = mt_cols.rename(columns={
        "Elliott_State":      "MT_State",
        "EW_Score":           "MT_Score",
        "Trade_Quality":      "MT_Quality",
        "Ideal_Entry_Price":  "MT_Entry",
        "Stop_Loss":          "MT_Stop",
        "Recommended_Action": "MT_Action",
        "Hold_Period":        "MT_Hold",
    })

    st_pick = [c for c in [
        "Ticker", "Elliott_State", "EW_Score", "Trade_Quality",
        "Recommended_Action", "Hold_Period"
    ] if c in df_short.columns]
    st_cols = df_short[st_pick].copy()
    st_cols = st_cols.rename(columns={
        "Elliott_State":      "ST_State",
        "EW_Score":           "ST_Score",
        "Trade_Quality":      "ST_Quality",
        "Recommended_Action": "ST_Action",
        "Hold_Period":        "ST_Hold",
    })

    merged = merged.merge(mt_cols, on="Ticker", how="inner")
    merged = merged.merge(st_cols, on="Ticker", how="inner")

    merged["Alignment_Type"] = merged.apply(
        lambda r: multi_tf_alignment_score(
            r.get("MN_State", "N/A"),
            r["LT_State"],
            r["MT_State"],
            r["ST_State"]
        ),
        axis=1
    )

    merged["Combined_Score"] = (
        merged["MN_Score"] * 0.2 +
        merged["LT_Score"] * 0.35 +
        merged["MT_Score"] * 0.25 +
        merged["ST_Score"] * 0.2
    ).round(1)

    _order = {
        "PLATINUM - Monthly+Weekly+Daily Bottom Aligned":   0,
        "GOLD - Monthly+Weekly Aligned, Daily Confirming":  1,
        "SILVER - Monthly+Daily Aligned":                   2,
        "STRONG - Weekly+Daily Aligned, No Conflict":       3,
        "GOOD - All 3 TF in Wave 3":                        4,
        "NEUTRAL - Wait for better alignment":              5,
        "LATE - Short-term extended, reduce size":          6,
        "HIGHER TF CONFLICT - SKIP":                        7,
        "MONTHLY CONFLICT - SKIP":                          8,
    }
    merged["_sort"] = merged["Alignment_Type"].map(_order).fillna(99)
    merged = merged.sort_values(["_sort", "Combined_Score"], ascending=[True, False])
    merged = merged.drop(columns=["_sort"])

    final_cols = [c for c in [
        "Ticker", "Company", "Sector", "SubIndustry",
        "Alignment_Type", "Combined_Score", "Current_Price",
        "MN_State", "MN_Score", "MN_Quality",
        "LT_State", "LT_Score", "LT_Quality", "LT_Entry", "LT_Stop", "LT_Action", "LT_Hold", "LT_Risk",
        "MT_State", "MT_Score", "MT_Quality", "MT_Entry", "MT_Stop", "MT_Action", "MT_Hold",
        "ST_State", "ST_Score", "ST_Quality", "ST_Action", "ST_Hold",
        "Extension_Target"
    ] if c in merged.columns]

    return merged[final_cols].reset_index(drop=True)

df_long = pd.DataFrame(results_lt)
df_mid = pd.DataFrame(results_mt)
df_short = pd.DataFrame(results_st)

if not df_long.empty:
    df_long = df_long[col_order].sort_values("EW_Score", ascending=False).reset_index(drop=True)
if not df_mid.empty:
    df_mid = df_mid[col_order].sort_values("EW_Score", ascending=False).reset_index(drop=True)
if not df_short.empty:
    df_short = df_short[col_order].sort_values("EW_Score", ascending=False).reset_index(drop=True)


# ────────────────────────────────────────────────────────────────
# ALL 3 TIMEFRAMES BULLISH
# ────────────────────────────────────────────────────────────────
def is_bullish_state(state_text):
    if pd.isna(state_text):
        return False

    s = str(state_text).upper()

    bullish_keywords = [
        "CAPITULATION","ACCUMULATION","WAVE 2 BUY ZONE","WAVE 3",
        "3RD OF 3RD","BREAKOUT","RIDE IT","ENTER NOW","BUY BREAKOUT",
        "BUY HERE","STARTING","ACTIVE"
    ]
    bearish_keywords = [
        "BEAR","BLOWOFF","EXIT","TRAP","REDUCE","LATE BULL",
        "SIDEWAYS","TRANSITIONING","WATCH LIST","SETUP FORMING",
        "WAIT FOR SIGNAL","STAND ASIDE"
    ]

    if any(k in s for k in bearish_keywords):
        return False
    return any(k in s for k in bullish_keywords)

if not df_long.empty and not df_mid.empty and not df_short.empty:
    lt_cols = [
        "Ticker","Company","Sector","SubIndustry",
        "EW_Score","Trade_Quality","Elliott_State","Current_Price",
        "Current_Wave_Target","Ideal_Entry_Price","Stop_Loss","Extension_Target",
        "Recommended_Action","Pivot_Low","Pivot_Low_Date","Pivot_Low_Price","Pivot_Low_Bars_Ago",
        "Pivot_High","Pivot_High_Date","Pivot_High_Price","Pivot_High_Bars_Ago"
    ]
    mt_cols = lt_cols.copy()
    st_cols = lt_cols.copy()

    df_lt_m = df_long[lt_cols].copy().rename(columns={
        "EW_Score": "LT_EW_Score",
        "Trade_Quality": "LT_Trade_Quality",
        "Elliott_State": "LT_State",
        "Current_Price": "LT_Price",
        "Current_Wave_Target": "LT_Target",
        "Ideal_Entry_Price": "LT_Entry",
        "Stop_Loss": "LT_Stop",
        "Extension_Target": "LT_Extension",
        "Recommended_Action": "LT_Action",
        "Pivot_Low": "LT_Pivot_Low",
        "Pivot_Low_Date": "LT_Pivot_Low_Date",
        "Pivot_Low_Price": "LT_Pivot_Low_Price",
        "Pivot_Low_Bars_Ago": "LT_Pivot_Low_Bars_Ago",
        "Pivot_High": "LT_Pivot_High",
        "Pivot_High_Date": "LT_Pivot_High_Date",
        "Pivot_High_Price": "LT_Pivot_High_Price",
        "Pivot_High_Bars_Ago": "LT_Pivot_High_Bars_Ago"
    })

    df_mt_m = df_mid[mt_cols].copy().rename(columns={
        "EW_Score": "MT_EW_Score",
        "Trade_Quality": "MT_Trade_Quality",
        "Elliott_State": "MT_State",
        "Current_Price": "MT_Price",
        "Current_Wave_Target": "MT_Target",
        "Ideal_Entry_Price": "MT_Entry",
        "Stop_Loss": "MT_Stop",
        "Extension_Target": "MT_Extension",
        "Recommended_Action": "MT_Action",
        "Pivot_Low": "MT_Pivot_Low",
        "Pivot_Low_Date": "MT_Pivot_Low_Date",
        "Pivot_Low_Price": "MT_Pivot_Low_Price",
        "Pivot_Low_Bars_Ago": "MT_Pivot_Low_Bars_Ago",
        "Pivot_High": "MT_Pivot_High",
        "Pivot_High_Date": "MT_Pivot_High_Date",
        "Pivot_High_Price": "MT_Pivot_High_Price",
        "Pivot_High_Bars_Ago": "MT_Pivot_High_Bars_Ago"
    })

    df_st_m = df_short[st_cols].copy().rename(columns={
        "EW_Score": "ST_EW_Score",
        "Trade_Quality": "ST_Trade_Quality",
        "Elliott_State": "ST_State",
        "Current_Price": "ST_Price",
        "Current_Wave_Target": "ST_Target",
        "Ideal_Entry_Price": "ST_Entry",
        "Stop_Loss": "ST_Stop",
        "Extension_Target": "ST_Extension",
        "Recommended_Action": "ST_Action",
        "Pivot_Low": "ST_Pivot_Low",
        "Pivot_Low_Date": "ST_Pivot_Low_Date",
        "Pivot_Low_Price": "ST_Pivot_Low_Price",
        "Pivot_Low_Bars_Ago": "ST_Pivot_Low_Bars_Ago",
        "Pivot_High": "ST_Pivot_High",
        "Pivot_High_Date": "ST_Pivot_High_Date",
        "Pivot_High_Price": "ST_Pivot_High_Price",
        "Pivot_High_Bars_Ago": "ST_Pivot_High_Bars_Ago"
    })

    df_all = (
        df_lt_m.merge(df_mt_m, on=["Ticker","Company","Sector","SubIndustry"], how="inner")
               .merge(df_st_m, on=["Ticker","Company","Sector","SubIndustry"], how="inner")
    )

    df_all["LT_Bull"] = df_all["LT_State"].apply(is_bullish_state)
    df_all["MT_Bull"] = df_all["MT_State"].apply(is_bullish_state)
    df_all["ST_Bull"] = df_all["ST_State"].apply(is_bullish_state)

    df_all_bull = df_all[(df_all["LT_Bull"]) & (df_all["MT_Bull"]) & (df_all["ST_Bull"])].copy()

    df_all_bull["Triple_Bull_Score"] = (
        df_all_bull["LT_EW_Score"] * 0.40 +
        df_all_bull["MT_EW_Score"] * 0.35 +
        df_all_bull["ST_EW_Score"] * 0.25
    ).round(2)

    def alignment_label(row):
        lt = str(row["LT_State"]).upper()
        mt = str(row["MT_State"]).upper()
        st = str(row["ST_State"]).upper()

        if "WAVE 3" in lt and "WAVE 3" in mt and ("3RD OF 3RD" in st or "BREAKOUT" in st or "WAVE 3" in st):
            return "PERFECT BULL ALIGNMENT"
        if (("ACCUMULATION" in lt) or ("WAVE 2" in lt)) and (("WAVE 3" in mt) or ("STARTING" in mt)) and (("BREAKOUT" in st) or ("WAVE 3" in st)):
            return "EARLY BULL ALIGNMENT"
        return "BULL ALIGNMENT"

    def master_action(row):
        if row["Triple_Bull_Score"] >= 88:
            return "STRONG BUY"
        elif row["Triple_Bull_Score"] >= 78:
            return "BUY / ADD"
        else:
            return "WATCH FOR ENTRY"

    def professor_note(row):
        return (
            f"Long-term trend: {row['LT_State']}. "
            f"Mid-term structure: {row['MT_State']}. "
            f"Short-term timing: {row['ST_State']}. "
            f"For a beginner, use the short-term or mid-term pullback entry instead of chasing strength."
        )

    df_all_bull["Alignment_Type"] = df_all_bull.apply(alignment_label, axis=1)
    df_all_bull["Master_Action"] = df_all_bull.apply(master_action, axis=1)
    df_all_bull["Professor_Note"] = df_all_bull.apply(professor_note, axis=1)

    df_all_bull = df_all_bull[
        [
            "Ticker","Company","Sector","SubIndustry","Triple_Bull_Score","Alignment_Type","Master_Action","Professor_Note",

            "LT_EW_Score","LT_Trade_Quality","LT_State","LT_Price","LT_Target","LT_Entry","LT_Stop","LT_Extension","LT_Action",
            "LT_Pivot_Low","LT_Pivot_Low_Date","LT_Pivot_Low_Price","LT_Pivot_Low_Bars_Ago",
            "LT_Pivot_High","LT_Pivot_High_Date","LT_Pivot_High_Price","LT_Pivot_High_Bars_Ago",

            "MT_EW_Score","MT_Trade_Quality","MT_State","MT_Price","MT_Target","MT_Entry","MT_Stop","MT_Extension","MT_Action",
            "MT_Pivot_Low","MT_Pivot_Low_Date","MT_Pivot_Low_Price","MT_Pivot_Low_Bars_Ago",
            "MT_Pivot_High","MT_Pivot_High_Date","MT_Pivot_High_Price","MT_Pivot_High_Bars_Ago",

            "ST_EW_Score","ST_Trade_Quality","ST_State","ST_Price","ST_Target","ST_Entry","ST_Stop","ST_Extension","ST_Action",
            "ST_Pivot_Low","ST_Pivot_Low_Date","ST_Pivot_Low_Price","ST_Pivot_Low_Bars_Ago",
            "ST_Pivot_High","ST_Pivot_High_Date","ST_Pivot_High_Price","ST_Pivot_High_Bars_Ago"
        ]
    ].sort_values(["Triple_Bull_Score","Ticker"], ascending=[False, True]).reset_index(drop=True)
else:
    df_all_bull = pd.DataFrame()


# Monthly timeframe DataFrame
df_monthly = pd.DataFrame(results_mn)
if not df_monthly.empty and "EW_Score" in df_monthly.columns:
    mn_col_order = [c for c in col_order if c in df_monthly.columns]
    df_monthly = df_monthly[mn_col_order].sort_values("EW_Score", ascending=False).reset_index(drop=True)

# Multi-timeframe alignment sheet
df_multi_tf = build_multi_tf_sheet(df_monthly, df_long, df_mid, df_short)


# ────────────────────────────────────────────────────────────────
# BUY LIST BUILDERS
# ────────────────────────────────────────────────────────────────
def safe_num(x):
    try:
        return float(x)
    except Exception:
        return np.nan

def build_buy_list(df, timeframe_label):
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()

    out["Current_Price_num"] = out["Current_Price"].apply(safe_num)
    out["Ideal_Entry_Price_num"] = out["Ideal_Entry_Price"].apply(safe_num)
    out["Stop_Loss_num"] = out["Stop_Loss"].apply(safe_num)
    out["Current_Wave_Target_num"] = out["Current_Wave_Target"].apply(safe_num)
    out["Extension_Target_num"] = out["Extension_Target"].apply(safe_num)

    out["Pct_to_Ideal_Entry"] = np.where(
        out["Ideal_Entry_Price_num"].notna() & (out["Ideal_Entry_Price_num"] > 0),
        ((out["Current_Price_num"] - out["Ideal_Entry_Price_num"]) / out["Ideal_Entry_Price_num"] * 100).round(2),
        np.nan
    )

    out["Pct_to_Stop"] = np.where(
        out["Stop_Loss_num"].notna() & (out["Current_Price_num"] > 0),
        ((out["Current_Price_num"] - out["Stop_Loss_num"]) / out["Current_Price_num"] * 100).round(2),
        np.nan
    )

    out["Pct_to_Target"] = np.where(
        out["Current_Wave_Target_num"].notna() & (out["Current_Price_num"] > 0),
        ((out["Current_Wave_Target_num"] - out["Current_Price_num"]) / out["Current_Price_num"] * 100).round(2),
        np.nan
    )

    out["Entry_Closeness_Score"] = np.where(
        out["Pct_to_Ideal_Entry"].notna(),
        100 - np.abs(out["Pct_to_Ideal_Entry"]),
        0
    )

    buy_states = [
        "WAVE 2", "WAVE 3", "ACCUMULATION", "CAPITULATION", "BREAKOUT", "3RD OF 3RD"
    ]
    buy_actions = [
        "BUY", "ENTER LONG", "BUY HERE", "BUY NOW / LONG", "BUY / BUILD POSITION", "HOLD / ADD DIPS"
    ]
    avoid_words = [
        "EXIT", "TRAP", "BLOWOFF", "BEAR", "REDUCE", "AVOID"
    ]

    def is_buy_candidate(row):
        state = str(row.get("Elliott_State", "")).upper()
        action = str(row.get("Recommended_Action", "")).upper()
        quality = str(row.get("Trade_Quality", "")).upper()

        bullish = any(x in state for x in buy_states) or any(x in action for x in buy_actions)
        bad = any(x in state for x in avoid_words) or any(x in action for x in avoid_words)
        decent_quality = any(x in quality for x in ["HIGH QUALITY", "VERY GOOD", "EARLY", "MEDIUM"])

        return bullish and (not bad) and decent_quality

    out = out[out.apply(is_buy_candidate, axis=1)].copy()

    out["Tomorrow_Buy_Rank"] = (
        out["EW_Score"].fillna(0) * 0.45 +
        out["Entry_Closeness_Score"].fillna(0) * 0.30 +
        out["Pct_to_Target"].clip(lower=0).fillna(0) * 0.15 +
        out["Pct_to_Stop"].clip(lower=0).fillna(0) * 0.10
    ).round(2)

    out["Timeframe"] = timeframe_label

    preferred_cols = [
        "Ticker","Company","Sector","SubIndustry","Fundamentally strong","Timeframe",
        "Tomorrow_Buy_Rank","EW_Score","Trade_Quality","Elliott_State",
        "Current_Price","Ideal_Entry_Price","Pct_to_Ideal_Entry",
        "Pullback_Buy_Zone","Stop_Loss","Pct_to_Stop",
        "Current_Wave_Target","Pct_to_Target","Extension_Target",
        "Recommended_Action","Beginner_Note","Key_Signals"
    ]

    preferred_cols = [c for c in preferred_cols if c in out.columns]
    out = out[preferred_cols].sort_values(
        ["Tomorrow_Buy_Rank","EW_Score","Ticker"],
        ascending=[False, False, True]
    ).reset_index(drop=True)

    return out

df_long_buys = build_buy_list(df_long, "LongTerm")
df_mid_buys = build_buy_list(df_mid, "MidTerm")
df_short_buys = build_buy_list(df_short, "ShortTerm")

df_tomorrow_buys = pd.concat(
    [df_short_buys, df_mid_buys, df_long_buys],
    ignore_index=True
).sort_values(
    ["Tomorrow_Buy_Rank","EW_Score","Ticker"],
    ascending=[False, False, True]
).reset_index(drop=True)


# ────────────────────────────────────────────────────────────────
# EXCEL HELPERS
# ────────────────────────────────────────────────────────────────
def thin_border():
    side = Side(style="thin", color="BDBDBD")
    return Border(left=side, right=side, top=side, bottom=side)

def score_fill(score):
    if score >= 88:
        return PatternFill("solid", fgColor="00C853")
    if score >= 78:
        return PatternFill("solid", fgColor="64DD17")
    if score >= 68:
        return PatternFill("solid", fgColor="FFD600")
    return PatternFill("solid", fgColor="FF6D00")

def state_fill(state):
    s = str(state).upper()
    if "3RD OF 3RD" in s or "EXPLOSIVE" in s:
        return PatternFill("solid", fgColor="00E5FF")
    if "WAVE 3" in s or "BREAKOUT" in s:
        return PatternFill("solid", fgColor="00C853")
    if "WAVE 2" in s or "BUY ZONE" in s:
        return PatternFill("solid", fgColor="FF9800")
    if "ACCUMULATION" in s or "CAPITULATION" in s:
        return PatternFill("solid", fgColor="B9F6CA")
    if "TRIANGLE" in s:
        return PatternFill("solid", fgColor="B0BEC5")
    if "BEAR" in s or "EXIT" in s or "TRAP" in s or "BLOWOFF" in s:
        return PatternFill("solid", fgColor="EF5350")
    return PatternFill("solid", fgColor="ECEFF1")

def action_fill(action_text):
    s = str(action_text).upper()
    if any(x in s for x in ["BUY","LONG","HOLD","ADD","BUILD"]):
        return PatternFill("solid", fgColor="00C853"), Font(bold=True, size=9, color="FFFFFF")
    if any(x in s for x in ["EXIT","AVOID","SELL","TRIM","SHORT","REDUCE"]):
        return PatternFill("solid", fgColor="EF5350"), Font(bold=True, size=9, color="FFFFFF")
    return PatternFill("solid", fgColor="FFD600"), Font(bold=True, size=9, color="000000")

def quality_fill(q):
    s = str(q).upper()
    if "VERY GOOD" in s or "HIGH QUALITY" in s:
        return PatternFill("solid", fgColor="00C853"), Font(bold=True, size=9, color="FFFFFF")
    if "EARLY" in s or "MEDIUM" in s:
        return PatternFill("solid", fgColor="FFD600"), Font(bold=True, size=9, color="000000")
    if "LOWER" in s or "AVOID" in s or "WAIT" in s:
        return PatternFill("solid", fgColor="FF6D00"), Font(bold=True, size=9, color="FFFFFF")
    return PatternFill("solid", fgColor="ECEFF1"), Font(bold=True, size=9, color="000000")


# ────────────────────────────────────────────────────────────────
# CHEAT SHEET TAB
# ────────────────────────────────────────────────────────────────
def add_cheat_sheet_tab(wb):
    ws = wb.create_sheet(title="Cheat_Sheet", index=0)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 62

    ws.merge_cells("A1:C1")
    ws["A1"] = "ELLIOTT WAVE CHEAT SHEET — HOW TO READ THIS WORKBOOK"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="263238")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    ws["A2"] = "Read in order: State → Quality → Entry → Stop → Target → Note"
    ws["A2"].font = Font(italic=True, size=10, color="37474F")
    ws["A2"].fill = PatternFill("solid", fgColor="ECEFF1")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 22

    def section(row, title, color="1565C0"):
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 22

    def headers(row):
        for col, text in zip(["A", "B", "C"], ["What you see", "What it means", "What you should do"]):
            c = ws[f"{col}{row}"]
            c.value = text
            c.font = Font(bold=True, size=10, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="455A64")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border()
        ws.row_dimensions[row].height = 22

    def add_row(row, a, b, c, fill="FFFFFF"):
        for col, val in zip(["A", "B", "C"], [a, b, c]):
            cell = ws[f"{col}{row}"]
            cell.value = val
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(size=10, color="000000")
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[row].height = 48

    section(4, "1) BEST ORDER TO READ A STOCK")
    headers(5)
    add_row(6, "Elliott_State", "Current phase of the stock cycle", "This is the first thing to read.")
    add_row(7, "Trade_Quality", "How clean or risky the setup is", "Prefer HIGH QUALITY or VERY GOOD first.")
    add_row(8, "Ideal_Entry_Price / Pullback_Buy_Zone", "Best entry area", "Do not chase far above these levels.")
    add_row(9, "Stop_Loss / Invalidation_Level", "Where the idea is wrong", "If price breaks this, exit and reassess.")
    add_row(10, "Current_Wave_Target / Extension_Target", "Likely target and stronger scenario", "Use as target zones, not guaranteed exact prices.")

    section(12, "2) HOW TO INTERPRET MAIN STATES", color="2E7D32")
    headers(13)
    add_row(14, "ACCUMULATION", "Early base building after weakness", "Good watchlist candidate; enter small or wait for more confirmation.", "F1F8E9")
    add_row(15, "CAPITULATION", "Panic may be ending", "High opportunity but high risk. Buy small, never all-in.", "F1F8E9")
    add_row(16, "WAVE 2 Buy Zone", "Pullback inside bullish structure", "Best beginner setup. Buy near entry zone, not far above it.", "FFF8E1")
    add_row(17, "WAVE 3 Starting", "Trend is beginning to accelerate", "Good setup. Enter on confirmation or a small pullback.", "E8F5E9")
    add_row(18, "WAVE 3 Active", "Strong trend already underway", "Great if already in. If new, wait for a pullback instead of chasing.", "E8F5E9")
    add_row(19, "3rd of 3rd Wave", "Explosive momentum", "Good for holders, risky for late buyers. Wait for pause/pullback.", "E0F7FA")
    add_row(20, "TRIANGLE", "Compression before possible thrust", "Wait for breakout confirmation; do not guess.", "ECEFF1")
    add_row(21, "WAVE 5", "Late-stage uptrend", "Avoid chasing. Better to protect profits or wait for correction.", "FFF3E0")
    add_row(22, "BLOWOFF / TRAP / BEAR", "Exhaustion or failed bullish structure", "Avoid fresh long entries.", "FFEBEE")

    section(24, "3) SIMPLE ACTION RULES", color="6A1B9A")
    headers(25)
    add_row(26, "BUY / Consider Entry", "State is bullish + quality is good + price near entry zone", "Best case: Wave 2 Buy Zone or Wave 3 Starting.")
    add_row(27, "WAIT", "Price is too far above ideal entry or setup is not confirmed", "Patience is a position too.")
    add_row(28, "AVOID", "Late wave, trap, blowoff, or bearish condition", "Preserve cash and move on.")
    add_row(29, "BEST TAB TO START", "All_3_Bullish", "Use this first because long, mid, and short timeframes are aligned.")

    section(31, "4) GOLDEN BEGINNER RULES", color="BF360C")
    headers(32)
    add_row(33, "Rule 1", "Buy pullbacks, not excitement", "If price is already far above entry, wait.")
    add_row(34, "Rule 2", "Respect invalidation, not hope", "If stop/invalidation breaks, exit.")
    add_row(35, "Rule 3", "Use targets as zones, not promises", "Take them seriously, but not literally.")
    add_row(36, "Rule 4", "Higher timeframe wins", "If short-term is bullish but long-term is weak, be cautious.")
    add_row(37, "Rule 5", "Start small", "Especially if you are new or the setup is early/uncertain.")

    ws.merge_cells("A39:C39")
    ws["A39"] = "Memory aid: Buy pullbacks, not excitement; respect invalidation, not hope."
    ws["A39"].font = Font(bold=True, italic=True, size=11, color="FFFFFF")
    ws["A39"].fill = PatternFill("solid", fgColor="37474F")
    ws["A39"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[39].height = 24

    ws.freeze_panes = "A5"


# ────────────────────────────────────────────────────────────────
# DATA SHEET WRITER
# ────────────────────────────────────────────────────────────────
def add_sheet_to_workbook(wb, df, sheet_name, theme_color, title_text):
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_view.showGridLines = False

    if df is None or df.empty:
        df = pd.DataFrame({"Message": [f"No rows available for {sheet_name}"]})

    max_col_letter = get_column_letter(len(df.columns) if len(df.columns) > 0 else 1)

    ws.merge_cells(f"A1:{max_col_letter}1")
    ws["A1"] = title_text
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=theme_color)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells(f"A2:{max_col_letter}2")
    ws["A2"] = "Elliott Wave + Fibonacci projection + beginner-friendly interpretation | Generated: " + datetime.today().strftime("%Y-%m-%d")
    ws["A2"].font = Font(italic=True, size=9, color="455A64")
    ws["A2"].fill = PatternFill("solid", fgColor="ECEFF1")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=3, column=ci, value=col.replace("_", " "))
        cell.font = Font(bold=True, size=9, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=theme_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[3].height = 30

    for ri, row in enumerate(df.itertuples(index=False), start=4):
        alt_fill = "F7F7F7" if ri % 2 == 0 else "FFFFFF"

        for ci, val in enumerate(row, start=1):
            col_name = df.columns[ci - 1]
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=alt_fill)
            cell.font = Font(size=9, color="000000")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

            if col_name == "Ticker":
                cell.font = Font(bold=True, size=11, color="0D47A1")
            elif col_name in ["Company", "Sector", "SubIndustry"]:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col_name in ["EW_Score", "Triple_Bull_Score", "LT_EW_Score", "MT_EW_Score", "ST_EW_Score", "Tomorrow_Buy_Rank"]:
                try:
                    cell.fill = score_fill(float(val))
                    cell.font = Font(bold=True, size=10, color="FFFFFF")
                except Exception:
                    pass
            elif "State" in col_name:
                cell.fill = state_fill(str(val))
                cell.font = Font(bold=True, size=9, color="000000")
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col_name in ["Trade_Quality", "LT_Trade_Quality", "MT_Trade_Quality", "ST_Trade_Quality", "Alignment_Type"]:
                fill, font = quality_fill(str(val))
                cell.fill = fill
                cell.font = font
            elif "Action" in col_name:
                fill, font = action_fill(str(val))
                cell.fill = fill
                cell.font = font
            elif col_name in ["Beginner_Note", "If_Extended_Then", "Professor_Note", "Key_Signals"]:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                cell.font = Font(size=8, italic=True, color="263238")
            elif "RSI" in col_name:
                try:
                    rv = float(val)
                    color = "EF5350" if rv > 70 else "00C853" if rv < 40 else "FF9800"
                    cell.font = Font(bold=True, size=9, color=color)
                except Exception:
                    pass
            elif "Vol_Ratio" in col_name:
                try:
                    vv = float(val)
                    color = "00C853" if vv > 1.5 else "FF6D00" if vv < 0.75 else "455A64"
                    cell.font = Font(bold=True, size=9, color=color)
                except Exception:
                    pass

        ws.row_dimensions[ri].height = 42

    default_widths = {
        "Ticker": 10,
        "Company": 26,
        "Sector": 18,
        "SubIndustry": 26,
        "Timeframe": 12,

        "EW_Score": 10,
        "Tomorrow_Buy_Rank": 14,
        "Trade_Quality": 20,
        "Elliott_State": 30,
        "Current_Price": 11,
        "Current_Wave_Target": 14,
        "Ideal_Entry_Price": 14,
        "Pct_to_Ideal_Entry": 14,
        "Pullback_Buy_Zone": 16,
        "Stop_Loss": 11,
        "Pct_to_Stop": 12,
        "Invalidation_Level": 14,
        "Extension_Target": 13,
        "Pct_to_Target": 12,
        "If_Extended_Then": 42,
        "Beginner_Note": 52,
        "RSI": 8,
        "Vol_Ratio": 10,
        "Dow_MA_Score": 12,
        "OBV_Divergence": 12,
        "MACD_Signal": 12,
        "Pct_from_52w_High": 14,
        "Pct_from_52w_Low": 14,
        "Key_Signals": 40,
        "52w_High": 10,
        "52w_Low": 10,
        "Fib_61.8%": 10,
        "Fib_38.2%": 10,
        "Fib_161.8%_Target": 14,
        "Recommended_Action": 20,
        "Hold_Period": 16,
        "Risk_Management": 36,
        "Rule1_Wave2_OK": 14,
        "W4_Overlap_Risk": 14,

        "Pivot_Low": 10,
        "Pivot_Low_Date": 14,
        "Pivot_Low_Price": 14,
        "Pivot_Low_Bars_Ago": 16,
        "Pivot_High": 10,
        "Pivot_High_Date": 14,
        "Pivot_High_Price": 14,
        "Pivot_High_Bars_Ago": 16,
        "State_Duration_Bars": 18,
        "State_Duration_Human": 22,
        "Next_EW_State": 36,
        "Next_State_ETA_Bars": 18,
        "Next_State_ETA_Human": 22,
        "Transition_Trigger": 48,

        "Triple_Bull_Score": 14,
        "Alignment_Type": 22,
        "Master_Action": 18,
        "Professor_Note": 60,

        "LT_EW_Score": 11, "LT_Trade_Quality": 18, "LT_State": 24, "LT_Price": 10, "LT_Target": 11, "LT_Entry": 11, "LT_Stop": 10, "LT_Extension": 12, "LT_Action": 16,
        "LT_Pivot_Low": 10, "LT_Pivot_Low_Date": 14, "LT_Pivot_Low_Price": 14, "LT_Pivot_Low_Bars_Ago": 16,
        "LT_Pivot_High": 10, "LT_Pivot_High_Date": 14, "LT_Pivot_High_Price": 14, "LT_Pivot_High_Bars_Ago": 16,

        "MT_EW_Score": 11, "MT_Trade_Quality": 18, "MT_State": 24, "MT_Price": 10, "MT_Target": 11, "MT_Entry": 11, "MT_Stop": 10, "MT_Extension": 12, "MT_Action": 16,
        "MT_Pivot_Low": 10, "MT_Pivot_Low_Date": 14, "MT_Pivot_Low_Price": 14, "MT_Pivot_Low_Bars_Ago": 16,
        "MT_Pivot_High": 10, "MT_Pivot_High_Date": 14, "MT_Pivot_High_Price": 14, "MT_Pivot_High_Bars_Ago": 16,

        "ST_EW_Score": 11, "ST_Trade_Quality": 18, "ST_State": 24, "ST_Price": 10, "ST_Target": 11, "ST_Entry": 11, "ST_Stop": 10, "ST_Extension": 12, "ST_Action": 16,
        "ST_Pivot_Low": 10, "ST_Pivot_Low_Date": 14, "ST_Pivot_Low_Price": 14, "ST_Pivot_Low_Bars_Ago": 16,
        "ST_Pivot_High": 10, "ST_Pivot_High_Date": 14, "ST_Pivot_High_Price": 14, "ST_Pivot_High_Bars_Ago": 16
    }

    for ci, col in enumerate(df.columns, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = default_widths.get(col, 15)

    ws.freeze_panes = "E4"


# ────────────────────────────────────────────────────────────────
# SAVE WORKBOOK
# ────────────────────────────────────────────────────────────────
def save_master_workbook(
    df_long, df_mid, df_short, df_all_bull, df_multi_tf, df_monthly,
    df_long_buys, df_mid_buys, df_short_buys, df_tomorrow_buys,
    filename="Elliott_Wave_SP500_Master_Workbook.xlsx"
):
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    add_cheat_sheet_tab(wb)

    add_sheet_to_workbook(
        wb, df_tomorrow_buys, "Tomorrow_Buys", "BF360C",
        f"BEST BUY CANDIDATES FOR TOMORROW | {UNIVERSE_NAME}"
    )
    add_sheet_to_workbook(
        wb, df_long_buys, "LongTerm_Buys", "1B5E20",
        f"LONG-TERM BUY CANDIDATES | {UNIVERSE_NAME}"
    )
    add_sheet_to_workbook(
        wb, df_mid_buys, "MidTerm_Buys", "0D47A1",
        f"MID-TERM BUY CANDIDATES | {UNIVERSE_NAME}"
    )
    add_sheet_to_workbook(
        wb, df_short_buys, "ShortTerm_Buys", "4A148C",
        f"SHORT-TERM BUY CANDIDATES | {UNIVERSE_NAME}"
    )

    add_sheet_to_workbook(
        wb, df_long, "LongTerm", "2E7D32",
        f"ELLIOTT WAVE — LONG-TERM SUPER CYCLE SCREENER | {UNIVERSE_NAME} | Weekly 5y"
    )
    add_sheet_to_workbook(
        wb, df_mid, "MidTerm", "1565C0",
        f"ELLIOTT WAVE — MID-TERM SWING TRADE SCREENER | {UNIVERSE_NAME} | Daily 1y"
    )
    add_sheet_to_workbook(
        wb, df_short, "ShortTerm", "6A1B9A",
        f"ELLIOTT WAVE — SHORT-TERM MOMENTUM SCREENER | {UNIVERSE_NAME} | Daily 90d"
    )
    add_sheet_to_workbook(
        wb, df_all_bull, "All_3_Bullish", "004D40",
        f"ELLIOTT WAVE — ALL 3 TIMEFRAMES BULLISH | {UNIVERSE_NAME}"
    )
    add_sheet_to_workbook(
        wb, df_monthly, "Monthly_EW", "37474F",
        f"ELLIOTT WAVE — MONTHLY SUPER-CYCLE SCREENER | {UNIVERSE_NAME} | Monthly 10yr"
    )
    add_sheet_to_workbook(
        wb, df_multi_tf, "Multi_TF_Alignment", "1A237E",
        f"MULTI-TIMEFRAME ALIGNMENT | {UNIVERSE_NAME} | Monthly+Weekly+Daily"
    )

    if "df_backtest" in dir() and not df_backtest.empty and "Message" not in df_backtest.columns:
        add_sheet_to_workbook(wb, df_backtest, "Backtest_Results", "4A148C",
            f"ELLIOTT WAVE — BACKTEST RESULTS | {UNIVERSE_NAME}")
    try:
        wb.save(filename)               
        print(f"\nSaved workbook: {filename}")
    except Exception as e:
        print(f"\n[CRITICAL ERROR] Failed to save Excel file to disk: {e}")                


# ────────────────────────────────────────────────────────────────
# FINAL SAVE
# ────────────────────────────────────────────────────────────────
# ────────────────────────────────────────────────────────────────

# BACKTEST MODULE — Walk-Forward Signal Quality
# ────────────────────────────────────────────────────────────────
ENABLE_BACKTEST = False   # Set False to skip (adds ~30s per ticker)
BACKTEST_FWD_BARS = 20
BACKTEST_WINDOWS  = 6

if ENABLE_BACKTEST:
    print("\nRunning walk-forward backtest...")
    df_backtest = run_backtest_summary(
        TICKERS,
        horizon="long",
        lookback_windows=BACKTEST_WINDOWS,
        fwd_bars=BACKTEST_FWD_BARS
    )
    if not df_backtest.empty:
        buy_signals = df_backtest[df_backtest["Is_Buy_Signal"] == True]
        hit_rate = (buy_signals["Hit"] == "✅").mean() * 100 if len(buy_signals) else 0
        print(f"Backtest done: {len(df_backtest)} signals | Buy hit rate: {hit_rate:.1f}%")
    else:
        df_backtest = pd.DataFrame({"Message": ["No backtest data"]})
else:
    df_backtest = pd.DataFrame({"Message": ["Backtest disabled"]})

save_master_workbook(
    df_long=df_long,
    df_mid=df_mid,
    df_short=df_short,
    df_all_bull=df_all_bull,
    df_multi_tf=df_multi_tf,
    df_monthly=df_monthly,
    df_long_buys=df_long_buys,
    df_mid_buys=df_mid_buys,
    df_short_buys=df_short_buys,
    df_tomorrow_buys=df_tomorrow_buys,
    filename=OUTPUT_FILENAME
)

print("\nDone.")
print(f"Workbook created: {OUTPUT_FILENAME}")
print("Tabs: Cheat_Sheet | Tomorrow_Buys | LongTerm_Buys | MidTerm_Buys | ShortTerm_Buys | LongTerm | MidTerm | ShortTerm | All_3_Bullish | Monthly_EW | Multi_TF_Alignment")
